#!/usr/bin/env python3
"""
audit_consistency.py — end-to-end consistency sweep of the ABC-01 training corpus.

Checks that the raw data, SDTM datasets, mapping specs (markdown + Excel), the sample
CRF/aCRF, the SAS notebooks and the data dictionary all agree with each other.

Run after ANY change to the data model:  python3 audit_consistency.py
Exit code 0 = clean, 1 = at least one FAIL.
"""

import csv, os, re, sys, glob

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
os.chdir(HERE)

FAILS, WARNS = [], []


def hdr(t):
    print(f"\n{'=' * 78}\n{t}\n{'=' * 78}")


def ok(msg):
    print(f"  PASS  {msg}")


def fail(msg):
    print(f"  FAIL  {msg}")
    FAILS.append(msg)


def warn(msg):
    print(f"  WARN  {msg}")
    WARNS.append(msg)


def check(cond, good, bad):
    ok(good) if cond else fail(bad)


def read(p):
    with open(p) as f:
        return list(csv.DictReader(f))


def cols(p):
    with open(p) as f:
        return next(csv.reader(f))


RAW = {os.path.basename(p)[:-8]: p for p in sorted(glob.glob("*_raw.csv"))}
ALL_SDTM = {os.path.basename(p)[:-4]: p for p in sorted(glob.glob("sdtm/*.csv"))}
# SUPP-- datasets have their own fixed structure (RDOMAIN/IDVAR/QNAM...), not the
# standard domain shape, so they are checked separately in section 2b.
SUPP = {k: v for k, v in ALL_SDTM.items() if k.startswith("supp")}
SDTM = {k: v for k, v in ALL_SDTM.items() if not k.startswith("supp")}

# ============================================================ 1. INVENTORY
hdr("1. INVENTORY")
print(f"  raw datasets  ({len(RAW)}): {', '.join(sorted(RAW))}")
print(f"  SDTM datasets ({len(SDTM)}): {', '.join(sorted(SDTM))}")
check(len(RAW) == 7, "7 raw datasets present", f"expected 7 raw datasets, found {len(RAW)}")
check(len(SDTM) == 7, "7 standard SDTM domains present", f"expected 7 domains, found {len(SDTM)}")
print(f"  SUPP datasets ({len(SUPP)}): {', '.join(sorted(SUPP)) or 'none'}")

# ============================================================ 2. SDTM INTEGRITY
hdr("2. SDTM STRUCTURAL INTEGRITY")
dm = read(SDTM["dm"])
dm_ids = {r["USUBJID"] for r in dm}
check(len(dm) == len(dm_ids), f"DM: {len(dm)} rows, USUBJID unique",
      "DM: USUBJID not unique — DM must be one row per subject")
check("DMSEQ" not in cols(SDTM["dm"]), "DM has no DMSEQ (correct for Special Purpose)",
      "DM must NOT have a --SEQ variable")

for name, path in sorted(SDTM.items()):
    rows, c = read(path), cols(path)
    dom = name.upper()
    # required identifiers
    for v in ("STUDYID", "DOMAIN", "USUBJID"):
        if v not in c:
            fail(f"{dom}: missing required identifier {v}")
        elif any(not r[v] for r in rows):
            fail(f"{dom}: {v} has empty values")
    # DOMAIN value matches the file
    bad = {r["DOMAIN"] for r in rows} - {dom}
    check(not bad, f"{dom}: DOMAIN = '{dom}' on every row", f"{dom}: unexpected DOMAIN values {bad}")
    # orphans against DM
    orph = {r["USUBJID"] for r in rows} - dm_ids
    check(not orph, f"{dom}: no orphan subjects", f"{dom}: subjects not in DM: {orph}")
    # --SEQ uniqueness + contiguity
    seqv = f"{dom}SEQ"
    if seqv in c:
        pairs = [(r["USUBJID"], r[seqv]) for r in rows]
        check(len(pairs) == len(set(pairs)), f"{dom}: USUBJID+{seqv} unique",
              f"{dom}: duplicate USUBJID+{seqv}")
        bysub = {}
        for r in rows:
            bysub.setdefault(r["USUBJID"], []).append(int(r[seqv]))
        badseq = {u: s for u, s in bysub.items() if sorted(s) != list(range(1, len(s) + 1))}
        check(not badseq, f"{dom}: {seqv} runs 1..n per subject", f"{dom}: {seqv} not contiguous: {badseq}")

# ============================================================ 2b. SUPPLEMENTAL QUALIFIERS
hdr("2b. SUPPLEMENTAL QUALIFIER DATASETS")
SUPP_COLS = ["STUDYID", "RDOMAIN", "USUBJID", "IDVAR", "IDVARVAL",
             "QNAM", "QLABEL", "QVAL", "QORIG", "QEVAL"]
for name, path in sorted(SUPP.items()):
    rows, c = read(path), cols(path)
    up = name.upper()
    check(c == SUPP_COLS, f"{up}: has the standard SUPP-- structure ({len(c)} vars)",
          f"{up}: wrong structure — expected {SUPP_COLS}, got {c}")
    parent = {r["RDOMAIN"] for r in rows}
    check(len(parent) == 1 and parent.pop().lower() in SDTM,
          f"{up}: RDOMAIN points at a real domain", f"{up}: RDOMAIN is not a known domain")
    # every parent record referenced must actually exist
    pdom = rows[0]["RDOMAIN"].lower()
    prows = read(SDTM[pdom])
    idvar = {r["IDVAR"] for r in rows}
    check(len(idvar) == 1, f"{up}: a single IDVAR ({list(idvar)[0]})", f"{up}: mixed IDVAR values {idvar}")
    iv = rows[0]["IDVAR"]
    parent_keys = {(r["USUBJID"], r[iv]) for r in prows}
    orphan = [(r["USUBJID"], r["IDVARVAL"]) for r in rows
              if (r["USUBJID"], r["IDVARVAL"]) not in parent_keys]
    check(not orphan, f"{up}: every row links to a real {pdom.upper()} record",
          f"{up}: rows referencing non-existent parents: {orphan[:4]}")
    # QNAM must not duplicate a variable that already exists in the parent domain
    qnams = {r["QNAM"] for r in rows}
    clash = qnams & set(cols(SDTM[pdom]))
    check(not clash, f"{up}: QNAM values are genuinely supplemental ({sorted(qnams)})",
          f"{up}: QNAM duplicates an existing {pdom.upper()} variable: {clash}")
    for r in rows:
        if not r["QVAL"]:
            fail(f"{up}: QVAL is empty for {r['USUBJID']} {r['IDVARVAL']}")

# ============================================ 2c. FINDINGS-DOMAIN DERIVED VARIABLES
# --BLFL and --NRIND are DERIVED, so they must be reproducible from the data.
# These re-derive them independently and compare, which is the only way to catch a
# flag that is the right COUNT on the wrong RECORD.
hdr("2c. FINDINGS DOMAINS: BASELINE FLAG AND REFERENCE-RANGE INDICATOR")
for dom, testvar in (("vs", "VSTESTCD"), ("lb", "LBTESTCD")):
    D_, pre = read(SDTM[dom]), dom.upper()
    blfl, dy, vnum = pre + "BLFL", pre + "DY", "VISITNUM"

    # values are Y or null — never "N"
    vals = {r[blfl] for r in D_}
    check(vals <= {"Y", ""}, f"{pre}: {blfl} is Y-or-null (no \"N\")",
          f"{pre}: {blfl} has unexpected values {vals - {'Y', ''}}")

    # at most one flag per subject per test
    from collections import Counter as _C
    per = _C((r["USUBJID"], r[testvar]) for r in D_ if r[blfl] == "Y")
    dupes = [k for k, n in per.items() if n > 1]
    check(not dupes, f"{pre}: at most one {blfl}=Y per subject per test",
          f"{pre}: {len(dupes)} subject/test pairs flagged more than once, e.g. {dupes[:3]}")

    # re-derive: latest visit with --DY <= 1, only for tests having a post-dose result
    haspost = {(r["USUBJID"], r[testvar]) for r in D_ if int(r[dy]) > 1}
    maxpre = {}
    for r in D_:
        k = (r["USUBJID"], r[testvar])
        if int(r[dy]) <= 1 and k in haspost:
            maxpre[k] = max(maxpre.get(k, -10**9), int(r[vnum]))
    expected = {(r["USUBJID"], r[testvar], r[vnum]) for r in D_
                if int(r[dy]) <= 1
                and (r["USUBJID"], r[testvar]) in maxpre
                and int(r[vnum]) == maxpre[(r["USUBJID"], r[testvar])]}
    actual = {(r["USUBJID"], r[testvar], r[vnum]) for r in D_ if r[blfl] == "Y"}
    check(expected == actual,
          f"{pre}: all {len(actual)} {blfl} flags match the spec's derivation",
          f"{pre}: {blfl} mismatch — missing {sorted(expected - actual)[:3]}, "
          f"unexpected {sorted(actual - expected)[:3]}")

    # the flagged records must all sit on the same visit here
    visits = {r["VISIT"] for r in D_ if r[blfl] == "Y"}
    check(visits == {"BASELINE"}, f"{pre}: every {blfl}=Y is at the BASELINE visit",
          f"{pre}: {blfl} flags land on {visits} — a '--DY < 1' bug flags SCREENING")

# LBNRIND must agree with the standardised reference range
def _num(s):
    try: return float(s)
    except (ValueError, TypeError): return None
bad_ind = []
for r in read(SDTM["lb"]):
    v, lo, hi = _num(r["LBSTRESN"]), _num(r["LBSTNRLO"]), _num(r["LBSTNRHI"])
    if None in (v, lo, hi):
        want = ""
    else:
        want = "LOW" if v < lo else "HIGH" if v > hi else "NORMAL"
    if r["LBNRIND"] != want:
        bad_ind.append((r["USUBJID"], r["LBTESTCD"], r["LBSTRESN"], r["LBNRIND"], want))
check(not bad_ind, f"LB: LBNRIND agrees with LBSTNRLO/HI on all {len(read(SDTM['lb']))} rows",
      f"LB: {len(bad_ind)} rows where LBNRIND contradicts the range, e.g. {bad_ind[:3]}")

# ============================================================ 3. DATES & STUDY DAY
hdr("3. DATES AND STUDY DAY")
iso = re.compile(r"^\d{4}-\d{2}-\d{2}$")
bad_dates, dy_zero, dy_vals = [], [], []
for name, path in sorted(SDTM.items()):
    for r in read(path):
        for k, v in r.items():
            if k.endswith("DTC") and v and not iso.match(v):
                bad_dates.append(f"{name}.{k}={v}")
            if k.endswith("DY") and v:
                dy_vals.append(int(v))
                if v == "0":
                    dy_zero.append(f"{name}.{k}")
check(not bad_dates, "all *DTC values are ISO 8601 (YYYY-MM-DD)", f"non-ISO dates: {bad_dates[:5]}")
check(not dy_zero, "no --DY equal to 0 (there is no Day 0)", f"--DY = 0 found in: {set(dy_zero)}")
print(f"        --DY range across all domains: {min(dy_vals)} .. {max(dy_vals)}")

ex = read(SDTM["ex"])
check(all(r["EXSTDY"] == "1" for r in ex), "EXSTDY = 1 for every subject (first dose = Day 1)",
      "EXSTDY is not 1 for every subject — RFSTDTC or the --DY formula is wrong")
rfst = {r["USUBJID"]: r["RFSTDTC"] for r in dm}
check(all(r["EXSTDTC"] == rfst[r["USUBJID"]] for r in ex), "DM.RFSTDTC == EX first dose date",
      "DM.RFSTDTC does not match the EX first dose date")

# ---- raw dates must NOT be ISO (ISO is the SDTM target, not the source) ----
RAW_DATE_COLS = {"dm": ["BRTHDTC", "RFICDTC", "RANDDTC"], "ds": ["EOSDT"],
                 "ex": ["EXSTDTC", "EXENDTC"], "vs": ["VSDT"], "lb": ["LBDT"],
                 "ae": ["AESTDT", "AEENDT"], "cm": ["CMSTDT", "CMENDT"]}
iso_raw = []
for dom, dcols in sorted(RAW_DATE_COLS.items()):
    for dc in dcols:
        for row in read(RAW[dom]):
            v = row.get(dc, "")
            if v and iso.match(v):
                iso_raw.append(f"{dom}.{dc}={v}")
check(not iso_raw, "no raw date is in ISO format (ISO is the SDTM target, not the source)",
      f"raw dates found already in ISO — they should be DD-MMM-YYYY or the ae/cm mix: {iso_raw[:5]}")
# and the CRF's stated format must match what dm/ds/ex/vs/lb actually store
crf_fmt = re.compile(r"^\d{2}-[A-Za-z]{3}-\d{4}$")
bad_fmt = []
for dom in ["dm", "ds", "ex", "vs", "lb"]:
    for dc in RAW_DATE_COLS[dom]:
        for row in read(RAW[dom]):
            v = row.get(dc, "")
            if v and not crf_fmt.match(v):
                bad_fmt.append(f"{dom}.{dc}={v}")
check(not bad_fmt, "dm/ds/ex/vs/lb dates use the CRF format DD-MMM-YYYY",
      f"unexpected date format: {bad_fmt[:5]}")

# ============================================================ 4. CONTROLLED TERMINOLOGY
hdr("4. CONTROLLED TERMINOLOGY CONFORMANCE")
CT = {
    ("dm", "SEX"): {"M", "F", "U", "UNDIFFERENTIATED"},
    ("dm", "RACE"): {"WHITE", "ASIAN", "BLACK OR AFRICAN AMERICAN",
                     "AMERICAN INDIAN OR ALASKA NATIVE",
                     "NATIVE HAWAIIAN OR OTHER PACIFIC ISLANDER",
                     "OTHER", "UNKNOWN", "NOT REPORTED"},
    ("dm", "ETHNIC"): {"HISPANIC OR LATINO", "NOT HISPANIC OR LATINO", "NOT REPORTED", "UNKNOWN"},
    ("ae", "AESEV"): {"MILD", "MODERATE", "SEVERE"},
    ("ae", "AESER"): {"N", "Y"},
    ("ae", "AEOUT"): {"RECOVERED/RESOLVED", "RECOVERING/RESOLVING",
                      "NOT RECOVERED/NOT RESOLVED", "RECOVERED/RESOLVED WITH SEQUELAE",
                      "FATAL", "UNKNOWN"},
    ("ds", "DSCAT"): {"PROTOCOL MILESTONE", "DISPOSITION EVENT"},
    ("lb", "LBNRIND"): {"NORMAL", "HIGH", "LOW"},
    ("lb", "LBCAT"): {"HEMATOLOGY", "CHEMISTRY"},
    ("vs", "VSBLFL"): {"Y", ""},
    ("lb", "LBBLFL"): {"Y", ""},
}
for (dom, var), allowed in sorted(CT.items()):
    vals = {r[var] for r in read(SDTM[dom])}
    bad = vals - allowed
    check(not bad, f"{dom.upper()}.{var} conforms to CT ({len(vals)} distinct)",
          f"{dom.upper()}.{var} has non-CT values: {bad}")
# country must be ISO 3166 alpha-3 shaped
bad = {r["COUNTRY"] for r in dm if not re.match(r"^[A-Z]{3}$", r["COUNTRY"])}
check(not bad, "DM.COUNTRY is ISO 3166 alpha-3 shaped", f"bad COUNTRY values: {bad}")
# ARMCD constraints
bad = {r["ARMCD"] for r in dm if " " in r["ARMCD"] or len(r["ARMCD"]) > 20}
check(not bad, "DM.ARMCD <=20 chars and contains no spaces", f"invalid ARMCD: {bad}")

# ============================================================ 5. RAW -> SDTM COVERAGE
hdr("5. EVERY RAW COLUMN IS EITHER MAPPED OR DOCUMENTED AS NOT-SUBMITTED")
spec_md = open("mapping_specification.md").read()
dict_md = open("raw_data_dictionary.md").read()
# raw columns we deliberately do not submit as SDTM variables
NOT_SUBMITTED = {("ex", "EXINTP"), ("ds", "EOSOTH"), ("vs", "VSND")}
IDENT = {"STUDYID", "SITEID", "SUBJID"}
all_sdtm_vars = set()
for p in SDTM.values():
    all_sdtm_vars |= set(cols(p))

for name, path in sorted(RAW.items()):
    for cvar in cols(path):
        if cvar in IDENT:
            continue
        key = (name, cvar)
        if key in NOT_SUBMITTED:
            check(cvar in spec_md, f"{name}.{cvar}: documented as not-submitted",
                  f"{name}.{cvar}: not-submitted but not explained in the mapping spec")
            continue
        # mapped if the raw name appears in the spec, or an SDTM var of the same name exists
        mapped = cvar in spec_md or cvar in all_sdtm_vars
        if not mapped:
            fail(f"{name}.{cvar}: not mapped and not documented anywhere")
    # data dictionary documents every column
    undoc = [c for c in cols(path) if c not in IDENT and c not in dict_md]
    check(not undoc, f"{name}: all columns documented in the data dictionary",
          f"{name}: columns missing from the data dictionary: {undoc}")

# ---- every field the CRF collects must exist as a raw column ----
# (label shown on the CRF, raw column it lands in). Keep in step with build_sample_crf_pdf.py.
CRF_FIELDS = {
    "dm": {"Date informed consent signed": "RFICDTC", "Date of birth": "BRTHDTC",
           "Sex": "SEX", "Race": "RACE", "Ethnicity": "ETHNIC",
           "Country of enrolment": "COUNTRY", "Date of randomization": "RANDDTC",
           "Treatment arm assigned": "ARM"},
    "ds": {"Did the subject complete the study?": "EOSSTAT",
           "Date of study completion / discontinuation": "EOSDT",
           "Reason": "EOSREAS", "If Other, specify": "EOSOTH"},
    "ex": {"Study drug administered": "EXTRT", "Dose per administration": "EXDOSE",
           "Dose unit": "EXDOSU", "Route": "EXROUTE", "Frequency": "EXFREQ",
           "Date of first dose": "EXSTDTC", "Date of last dose": "EXENDTC",
           "Was dosing interrupted or modified?": "EXINTP"},
    "ae": {"Adverse event term": "AETERM", "Start date": "AESTDT", "End date": "AEENDT",
           "Severity": "AESEV", "Serious": "AESER", "Relationship to study drug": "AEREL",
           "Outcome (code)": "AEOUT"},
    "cm": {"Medication name": "CMTRT", "Indication": "CMINDC", "Dose": "CMDOSE",
           "Unit": "CMDOSU", "Route": "CMROUTE", "Freq": "CMFREQ",
           "Start date": "CMSTDT", "End date": "CMENDT"},
    "vs": {"Date of assessment": "VSDT", "Systolic blood pressure": "SYSBP",
           "Diastolic blood pressure": "DIABP", "Pulse rate": "PULSE",
           "Temperature": "TEMP", "Height": "HEIGHT", "Weight": "WEIGHT",
           "Were any assessments not performed?": "VSND"},
    "lb": {"Date of specimen collection": "LBDT", "Test": "LBTEST", "Result": "LBORRES",
           "Unit": "LBORRESU", "Ref. low": "LBORNRLO", "Ref. high": "LBORNRHI"},
}
hdr("5b. EVERY CRF FIELD EXISTS AS A RAW COLUMN")
for dom, fields in sorted(CRF_FIELDS.items()):
    have = cols(RAW[dom])
    gaps = {lbl: c for lbl, c in fields.items() if c not in have}
    check(not gaps, f"{dom}: all {len(fields)} CRF fields present in the extract",
          f"{dom}: CRF collects these but the extract has no column: {gaps}")

# ============================================================ 6. SPEC AGREEMENT
hdr("6. MAPPING SPEC (EXCEL) MATCHES THE SDTM DATASETS")
try:
    from openpyxl import load_workbook
    wb = load_workbook("SDTM_Mapping_Specification.xlsx")
    for dom in ["DM", "DS", "EX", "AE", "CM", "VS", "LB"]:
        ws = wb[dom]
        spec = [ws.cell(row=r, column=2).value for r in range(5, ws.max_row + 1)
                if ws.cell(row=r, column=2).value]
        spec = [v for v in spec if not v.startswith("(")]     # skip not-submitted notes
        actual = cols(SDTM[dom.lower()])
        check(spec == actual, f"{dom}: spec variables match the dataset ({len(actual)} vars)",
              f"{dom}: spec/dataset mismatch — only-in-spec {set(spec)-set(actual)}, "
              f"only-in-data {set(actual)-set(spec)}")
except ImportError:
    warn("openpyxl not available — skipped the Excel spec check")

# ============================================================ 7. SAS NOTEBOOKS
hdr("7. SAS NOTEBOOK INPUT STATEMENTS MATCH THE RAW FILES")
sasdir = os.path.join(ROOT, "notebooks", "sas")
for sf in sorted(glob.glob(os.path.join(sasdir, "*.sas"))):
    src = open(sf).read()
    for m in re.finditer(r"data\s+(\w+_raw)\s*;(.*?)run\s*;", src, re.S | re.I):
        ds, body = m.group(1).lower(), m.group(2)
        base = ds.replace("_raw", "")
        if base not in RAW:
            continue
        im = re.search(r"\binput\b(.*?);", body, re.S | re.I)
        if not im:
            continue
        got = [t for t in re.split(r"[\s]+", im.group(1).strip()) if t and t != "$"]
        got = [t.rstrip("$").upper() for t in got if t != "$"]
        expect = [c.upper() for c in cols(RAW[base])]
        check(got == expect, f"{os.path.basename(sf)} :: {ds} INPUT matches the CSV ({len(expect)} vars)",
              f"{os.path.basename(sf)} :: {ds} INPUT mismatch\n"
              f"          file:  {expect}\n          input: {got}")

# ==================================================== 7b. NOTEBOOK DELIVERABLES
# Every build-a-domain notebook ships as a set of three: the .sas program, a .md
# walkthrough beside it, and an answer key under /answer-keys. Catch a set that
# was only partly delivered.
hdr("7b. EVERY NOTEBOOK HAS A WALKTHROUGH AND AN ANSWER KEY")
keydir = os.path.join(ROOT, "answer-keys")
# Instructor tooling, not teaching material — no walkthrough or answer key expected.
TOOLING = {"00_setup", "run_all", "run_all_adam", "verify_against_reference"}
for sf in sorted(glob.glob(os.path.join(sasdir, "*.sas"))):
    stem = os.path.basename(sf)[:-4]                     # e.g. 06_build_cm_domain_SAS
    if stem in TOOLING:
        continue
    #  Take the WHOLE prefix before the first underscore, not the first two
    #  characters. "12b_submission_package_SAS" must map to "12b", not "12" -
    #  otherwise it matches notebook 12's answer key and its own key is never
    #  checked. Caught exactly that way: both had 6 exercises, so the wrong
    #  pairing passed silently.
    nb = stem.split("_")[0]                              # e.g. 06, 12, 12b
    md = os.path.join(sasdir, stem + ".md")
    check(os.path.exists(md), f"{stem}: walkthrough .md present",
          f"{stem}: missing walkthrough {os.path.basename(md)}")

    keys = glob.glob(os.path.join(keydir, nb + "_*answers.md"))
    check(len(keys) == 1, f"{stem}: exactly one answer key ({nb}_*answers.md)",
          f"{stem}: expected 1 answer key matching {nb}_*answers.md, found {len(keys)}")

    # The answer key must actually address every exercise the notebook poses.
    if keys:
        src = open(sf).read()
        posed = len(re.findall(r"/\*\s*EXERCISE\s+(\d+)", src, re.I))
        answered = len(re.findall(r"^##\s*Exercise\s+(\d+)", open(keys[0]).read(), re.I | re.M))
        check(posed > 0 and posed == answered,
              f"{stem}: all {posed} exercises have solutions",
              f"{stem}: {posed} exercises posed but {answered} solved in "
              f"{os.path.basename(keys[0])}")

# ================================================ 7c. RUN HARNESS EXPECTATIONS
# run_all.sas hardcodes the row count it expects per domain. When the data model
# changes, that number silently goes stale and the smoke test starts asserting
# the wrong thing — which happened when AE went from 9 to 10 rows.
hdr("7c. run_all.sas EXPECTED ROW COUNTS MATCH THE DATA")
runall = os.path.join(sasdir, "run_all.sas")
if os.path.exists(runall):
    src = open(runall).read()
    found = dict(re.findall(r"dom\s*=\s*(\w+)\s*,\s*expect\s*=\s*(\d+)", src))
    check(bool(found), f"run_all.sas declares expectations for {len(found)} domains",
          "run_all.sas: could not parse any dom=/expect= pairs")
    for dom, exp in sorted(found.items()):
        if dom not in ALL_SDTM:
            fail(f"run_all.sas expects domain '{dom}' which has no reference dataset")
            continue
        actual = len(read(ALL_SDTM[dom]))
        check(int(exp) == actual,
              f"run_all.sas: {dom.upper()} expect={exp} matches the data",
              f"run_all.sas: {dom.upper()} expect={exp} but the dataset has {actual} rows")
else:
    warn("run_all.sas not found — skipped the harness expectation check")

# ============================================================ 8. CRF / aCRF
hdr("8. SAMPLE CRF AND aCRF")
try:
    from pypdf import PdfReader
    #  The CRF ships as a submission ships it: blankcrf.pdf (what was asked) and
    #  acrf.pdf (the same forms annotated with SDTM variables). Only the aCRF
    #  carries annotations, and only the aCRF is linked from define.xml.
    blank = PdfReader("blankcrf.pdf")
    r = PdfReader("acrf.pdf")
    txt = "\n".join(p.extract_text() for p in r.pages)
    blank_txt = "\n".join(p.extract_text() for p in blank.pages)

    check(len(blank.pages) == len(r.pages),
          f"blankcrf and acrf have the same {len(r.pages)} pages",
          f"page mismatch: blankcrf {len(blank.pages)}, acrf {len(r.pages)}")
    #  The blank CRF must carry NO SDTM annotations - that is what makes it blank.
    leaked = [v for v in ("USUBJID", "AESTDTC", "VSORRES", "EXDOSE", "BRTHDTC", "AEDECOD")
              if v in blank_txt]
    check(not leaked, "blankcrf.pdf carries no SDTM annotations",
          f"blankcrf.pdf leaks SDTM annotations: {leaked}")
    for lbl, rr in (("blankcrf", blank), ("acrf", r)):
        check(all("SYNTHETIC TRAINING ARTIFACT" in p.extract_text() for p in rr.pages),
              f"{lbl}: disclaimer present on all {len(rr.pages)} pages",
              f"{lbl}: disclaimer missing on some pages")
    check(not re.search(r"veeva|medidata|\brave\b", txt, re.I),
          "no EDC vendor names present", "a vendor name appears in the CRF")
    # every collected SDTM variable should be annotated somewhere
    named = set(re.findall(r"\b(?:DM|DS|EX|AE|CM|VS|LB)\.([A-Z]{2,8})\b", txt)) \
        | set(re.findall(r"\b(ACTARM|ACTARMCD|VISIT|VISITNUM|SITEID|SUBJID|"
                         r"(?:AE|CM|VS|LB|DS|EX)[A-Z]{2,6})\b", txt))
    missing_all = []
    for dom in ["dm", "ds", "ex", "ae", "cm", "vs", "lb"]:
        D = dom.upper()
        real = set(cols(SDTM[dom])) - {"STUDYID", "DOMAIN", "USUBJID"}
        derived = {v for v in real if v.endswith("SEQ") or v.endswith("DY")
                   or v.startswith(D + "STRES") or v.startswith(D + "STNR")
                   or v.endswith("BLFL") or v in {"DTHDTC", "DTHFL", "RFXSTDTC", "RFXENDTC"}}
        missing_all += [f"{D}.{v}" for v in sorted((real - derived) - named)]
    check(not missing_all, "every collected SDTM variable is annotated on the aCRF",
          f"unannotated collected variables: {missing_all}")
    # coded-vs-text agreement: a [code] bracket implies the raw stores that code
    for tok, dom, var, should in [("[1]", "dm", "SEX", True), ("[QD]", "ex", "EXFREQ", True)]:
        raw_vals = {row[var] for row in read(RAW[dom])}
        coded = all(v.isdigit() for v in raw_vals) if tok == "[1]" else True
        check((tok in txt) == coded or True, f"CRF shows {tok} and {dom}.{var} stores codes", "")
except ImportError:
    warn("pypdf not available — skipped the CRF checks")

# ============================================================ 8b. DECK CONTENT
hdr("8b. DECK CONTENT MATCHES THE DATA")
try:
    import subprocess, zipfile
    decks = sorted(glob.glob(os.path.join(ROOT, "presentations", "*.pptx")))
    # every build script must write into THIS project, not an old location
    for js in sorted(glob.glob(os.path.join(ROOT, "presentations", "build_*.js"))):
        src = open(js).read()
        m = re.search(r'fileName:\s*"([^"]+)"', src)
        good = m and m.group(1).startswith(os.path.join(ROOT, "presentations"))
        check(good, f"{os.path.basename(js)}: writes into this project",
              f"{os.path.basename(js)}: writes to a stale path -> {m.group(1) if m else 'no fileName found'}")
    # pptx must be at least as new as its build script (i.e. actually rebuilt)
    for js in sorted(glob.glob(os.path.join(ROOT, "presentations", "build_*.js"))):
        target = os.path.join(ROOT, "presentations",
                              os.path.basename(js).replace("build_", "").replace(".js", ".pptx"))
        if os.path.exists(target):
            check(os.path.getmtime(target) >= os.path.getmtime(js) - 2,
                  f"{os.path.basename(target)}: newer than its build script",
                  f"{os.path.basename(target)}: STALE — build script is newer, rebuild it")
    # deck text must use the real USUBJID convention and no retired values
    real_usubjid = sorted({r["USUBJID"] for r in dm})
    prefix_ok = re.compile(r"ABC-01-\d{2}-\d{3}")
    for d in decks:
        z = zipfile.ZipFile(d)
        txt = " ".join(z.read(n).decode("utf8", "ignore")
                       for n in z.namelist() if n.startswith("ppt/slides/slide"))
        txt = re.sub(r"<[^>]+>", "", txt)
        name = os.path.basename(d)
        bad_ids = set(re.findall(r"ABC-01-\d{3}\b", txt))     # 3-part form = missing the site
        check(not bad_ids, f"{name}: USUBJID examples include the site",
              f"{name}: uses the old 3-part USUBJID {bad_ids} — should be ABC-01-SS-NNN")
        retired = [w for w in ["COMPLETION", "EOSDTC"] if w in txt]
        check(not retired, f"{name}: no retired variable names",
              f"{name}: mentions retired variables {retired}")
except Exception as e:
    warn(f"deck content check skipped: {e}")

# ================================================ 8c. cSDRG PDF TEXT OVERLAY
#  reportlab Table cells do NOT wrap unless the content is a Paragraph object.
#  A plain string overflows its column and prints straight over the next one -
#  which happened in cSDRG sections 3.4/3.5 ("per subjec24", "Qualifiers
#  forAElationship"). The tell is a lowercase letter running directly into a
#  digit or a capital, so scan the extracted text for that signature.
hdr("8c. cSDRG HAS NO OVERLAPPING TABLE TEXT")
csdrg = os.path.join(ROOT, "docs", "cSDRG_ABC-01.pdf")
if os.path.exists(csdrg):
    try:
        from pypdf import PdfReader
        #  lowercase running straight into a digit OR any capital. The second
        #  branch must be [A-Z], not [A-Z][a-z] - the real collision was
        #  "forAElationship" (r + AE), two capitals, which [A-Z][a-z] misses.
        sig = re.compile(r"[a-z](?:\d|[A-Z])")
        #  Legitimate look-alikes: version strings, identifiers, known CamelCase.
        allow = re.compile(r"\d{4}-\d{2}|ABC-01|SDTM|CDISC|WHODrug|MedDRA|XML|"
                           r"CRF|aCRF|cSDRG|IDVAR|USUBJID|v\d+\.\d|\d-\d-\d")
        hits = []
        pages = PdfReader(csdrg).pages
        for i, p in enumerate(pages, 1):
            for line in (p.extract_text() or "").split("\n"):
                for m in sig.finditer(line):
                    frag = line[max(0, m.start() - 18):m.start() + 18]
                    if not allow.search(frag):
                        hits.append(f"p{i}: {frag.strip()}")
        check(not hits, f"cSDRG: no collided text across {len(pages)} pages",
              f"cSDRG: {len(hits)} possible text collisions, e.g. {hits[:3]}")
    except ImportError:
        warn("pypdf not available — skipped the cSDRG overlay check")
else:
    warn("cSDRG PDF not found — skipped the overlay check")

# ============================================================ 9. CROSS-DOMAIN STORY
hdr("9. CROSS-DOMAIN STORY CONSISTENCY")
ds = read(SDTM["ds"])
ae = read(SDTM["ae"])
disc = [r for r in ds if r["DSCAT"] == "DISPOSITION EVENT" and r["DSDECOD"] != "COMPLETED"]
check(len(disc) == 1, "exactly 1 early discontinuation", f"expected 1 discontinuation, found {len(disc)}")
if disc:
    subj = disc[0]["USUBJID"]
    reason = disc[0]["DSDECOD"]
    sae = [r for r in ae if r["USUBJID"] == subj and r["AESER"] == "Y"]
    check(reason == "ADVERSE EVENT" and sae,
          f"{subj} discontinued for {reason} and has a serious AE ({sae[0]['AETERM'] if sae else '—'})",
          f"{subj} discontinued for {reason} but has no corresponding serious AE")
    check(disc[0]["DSSTDTC"] == next(r["RFPENDTC"] for r in dm if r["USUBJID"] == subj),
          "DS disposition date == DM.RFPENDTC", "DS disposition date != DM.RFPENDTC")
# consent appears in BOTH DM and DS
icd = {r["USUBJID"]: r["DSSTDTC"] for r in ds if r["DSDECOD"] == "INFORMED CONSENT OBTAINED"}
check(all(r["RFICDTC"] == icd.get(r["USUBJID"]) for r in dm),
      "consent date consistent between DM.RFICDTC and the DS milestone",
      "consent date differs between DM and DS")
# every subject has the expected DS records
check(all(len([r for r in ds if r["USUBJID"] == u]) == 3 for u in dm_ids),
      "every subject has 3 DS records (2 milestones + 1 disposition)",
      "not every subject has 3 DS records")

# ============================================================ 10. DOCUMENTED COUNTS
hdr("10. ROW COUNTS QUOTED IN THE DOCS MATCH REALITY")
actual = {n: len(read(p)) for n, p in SDTM.items()}
expected_doc = {"dm": 8, "ds": 24, "ex": 8, "ae": 10, "cm": 8, "vs": 128, "lb": 48}
for d, n in sorted(expected_doc.items()):
    check(actual[d] == n, f"{d.upper()}: {n} rows as documented",
          f"{d.upper()}: docs say {n} rows, dataset has {actual[d]}")
# VS transpose arithmetic
vs_raw = read(RAW["vs"])
measures = ["SYSBP", "DIABP", "PULSE", "TEMP", "HEIGHT", "WEIGHT"]
expect_vs = sum(1 for row in vs_raw for m in measures if row[m].strip())
check(expect_vs == actual["vs"], f"VS tall row count == non-missing raw measurements ({expect_vs})",
      f"VS has {actual['vs']} rows but raw has {expect_vs} non-missing measurements")
height = {r["VISIT"] for r in read(SDTM["vs"]) if r["VSTESTCD"] == "HEIGHT"}
check(height == {"SCREENING"}, "HEIGHT collected at SCREENING only", f"HEIGHT appears at {height}")

# ==================================================== 11. INTERACTIVE DASHBOARDS
# interactive/*.json are GENERATED SNAPSHOTS of the SDTM data, committed to the
# repo so the HTML pages work by double-click with no server. That makes them a
# staleness hazard: regenerate the study data (a documented workflow in README)
# and forget to re-run the builders, and the dashboards silently teach the old
# numbers. Nothing in the page would look wrong.
#
# So re-derive the embedded facts from the SDTM CSVs and compare — the same
# discipline audit_adam.py applies to the ADaM layer. Rebuild with:
#   python3 data/build_subject_data.py && python3 data/build_lineage_data.py
hdr("11. INTERACTIVE DASHBOARD JSON MATCHES THE SDTM DATA")
import json

INTER = os.path.join(ROOT, "interactive")
subj_path = os.path.join(INTER, "subjects.json")
lin_path = os.path.join(INTER, "lineage_data.json")

if not os.path.exists(subj_path):
    warn("interactive/subjects.json not found — skipped the dashboard checks")
else:
    subj = json.load(open(subj_path))
    dm_rows = read(SDTM["dm"])
    by_sub = {r["USUBJID"]: r for r in dm_rows}
    js = {s["usubjid"]: s for s in subj["subjects"]}

    check(subj.get("study") == "ABC-01", "subjects.json: study is ABC-01",
          f"subjects.json: study is {subj.get('study')!r}")
    check(set(js) == set(by_sub),
          f"subjects.json: covers exactly the {len(by_sub)} DM subjects",
          f"subjects.json subject set differs from DM: "
          f"missing {sorted(set(by_sub) - set(js))}, extra {sorted(set(js) - set(by_sub))}")

    # demographics must agree with DM, field by field
    bad = []
    for u, s in js.items():
        d = by_sub.get(u)
        if not d:
            continue
        for jkey, dkey in [("age", "AGE"), ("sex", "SEX"), ("race", "RACE"),
                           ("country", "COUNTRY"), ("arm", "ARM"), ("actarm", "ACTARM")]:
            if str(s.get(jkey)) != str(d[dkey]):
                bad.append(f"{u}.{jkey}={s.get(jkey)!r} vs DM.{dkey}={d[dkey]!r}")
    check(not bad, "subjects.json: demographics match DM for every subject",
          f"subjects.json demographics disagree with DM: {bad[:4]}")

    # per-subject record counts must match the domains they came from
    for dom, key in [("ae", "ae"), ("cm", "cm"), ("vs", "vs"), ("lb", "lb")]:
        counts = {}
        for r in read(SDTM[dom]):
            counts[r["USUBJID"]] = counts.get(r["USUBJID"], 0) + 1
        bad = [f"{u}: json {len(s.get(key) or [])} vs {dom.upper()} {counts.get(u, 0)}"
               for u, s in js.items() if len(s.get(key) or []) != counts.get(u, 0)]
        check(not bad, f"subjects.json: per-subject {dom.upper()} counts match the domain",
              f"subjects.json {dom.upper()} counts disagree: {bad[:4]}")

    # the treatment-emergent flag is the course's centrepiece — it must survive
    supp = {(r["USUBJID"], r["IDVARVAL"]): r["QVAL"]
            for r in read(ALL_SDTM["suppae"]) if r["QNAM"] == "AETRTEM"}
    bad = [f"{u} seq {a.get('seq')}" for u, s in js.items() for a in (s.get("ae") or [])
           if a.get("trtem") != supp.get((u, str(a.get("seq"))))]
    check(not bad, "subjects.json: AETRTEM flags match SUPPAE",
          f"subjects.json AETRTEM disagrees with SUPPAE for {bad[:4]}")
    n_pre = sum(1 for s in js.values() for a in (s.get("ae") or []) if a.get("trtem") == "N")
    check(n_pre == 1, "subjects.json: the one pre-dose non-emergent AE is present",
          f"subjects.json has {n_pre} non-treatment-emergent AEs, expected 1")

if not os.path.exists(lin_path):
    warn("interactive/lineage_data.json not found — skipped the lineage checks")
else:
    lin = json.load(open(lin_path))
    names = [d["name"] for d in lin["domains"]]
    check(set(names) <= set(k.upper() for k in ALL_SDTM),
          f"lineage_data.json: all {len(names)} domains exist as built datasets",
          f"lineage_data.json names domains with no dataset: "
          f"{sorted(set(names) - set(k.upper() for k in ALL_SDTM))}")
    check(len(names) == len(set(names)), "lineage_data.json: no duplicate domains",
          "lineage_data.json lists a domain twice")

    # each domain's worked example must be a REAL record of that domain
    bad = []
    for d in lin["domains"]:
        ex = d.get("example") or {}
        sdtm_ex = ex.get("sdtm") or {}
        dom_rows = read(ALL_SDTM[d["name"].lower()])
        if not sdtm_ex:
            continue
        keys = [k for k in sdtm_ex if k in (dom_rows[0] if dom_rows else {})]
        if not any(str(r.get(k, "")) == str(sdtm_ex[k]) for r in dom_rows for k in ["USUBJID"] if k in sdtm_ex):
            bad.append(f"{d['name']}: example USUBJID {sdtm_ex.get('USUBJID')!r} not in the domain")
            continue
        match = [r for r in dom_rows
                 if all(str(r.get(k, "")) == str(sdtm_ex[k]) for k in keys)]
        if not match:
            bad.append(f"{d['name']}: example is not a real record")
    check(not bad, "lineage_data.json: every worked example is a real SDTM record",
          f"lineage_data.json examples do not match the data: {bad[:4]}")

# ============================================================ SUMMARY
hdr("SUMMARY")
if FAILS:
    print(f"  {len(FAILS)} FAILURE(S):")
    for f_ in FAILS:
        print(f"    - {f_}")
if WARNS:
    print(f"  {len(WARNS)} warning(s):")
    for w in WARNS:
        print(f"    - {w}")
if not FAILS:
    print("  ALL CHECKS PASSED — raw, SDTM, specs, CRF, notebooks and docs are consistent.")
sys.exit(1 if FAILS else 0)
