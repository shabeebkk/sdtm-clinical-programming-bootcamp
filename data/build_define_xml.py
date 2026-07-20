#!/usr/bin/env python3
"""
build_define_xml.py — generate define.xml (Define-XML v2.0) for study ABC-01.

WHY THIS IS A SCRIPT AND NOT A HAND-WRITTEN FILE
In real work nobody types define.xml. It is GENERATED from the metadata
specification — the same spreadsheet the programmers map from. That is the whole
point of holding metadata as data: one source, two consumers (the programmers and
the define). This script does exactly that, reading SDTM_Mapping_Specification.xlsx.

Change the spec, re-run this, and the define follows. If they ever disagree, the
spec wins — which is the correct direction.

⚠️  SCOPE / HONESTY
This produces a structurally correct, well-formed Define-XML v2.0 file suitable
for TEACHING. It is NOT certified submission-grade: validating against the official
CDISC schema (define2-0-0.xsd) requires the XSD and a schema validator such as
Pinnacle 21, which is the tool a real submission would use. The script checks that
the XML parses and that every dataset and variable in the spec made it in.

Run (after build_mapping_spec_xlsx.py):  python3 build_define_xml.py
Writes: define/define.xml
"""

import os
import re
from xml.etree import ElementTree as ET
from xml.dom import minidom

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "define")
os.makedirs(OUT, exist_ok=True)

STUDYID = "ABC-01"
STUDYNAME = "A Phase 2 Study of Drug A versus Placebo"
STUDYDESC = ("Synthetic training study. Randomised, double-blind, placebo-controlled, "
             "8 subjects at 2 sites, 4 weeks of treatment.")
PROTOCOL = "ABC-01"
DEFINE_VERSION = "2.0.0"
STANDARD = "SDTM-IG"
STANDARD_VERSION = "3.3"

# Namespaces --------------------------------------------------------------
ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"
DEF_NS = "http://www.cdisc.org/ns/def/v2.0"
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
XLINK_NS = "http://www.w3.org/1999/xlink"
ET.register_namespace("", ODM_NS)
ET.register_namespace("def", DEF_NS)
ET.register_namespace("xsi", XSI_NS)
ET.register_namespace("xlink", XLINK_NS)


def q(ns, tag):
    return f"{{{ns}}}{tag}"


# Domain-level metadata: structure and keys. Keys matter — a reviewer uses them
# to know what makes a record unique.
DOMAIN_META = {
    "DM":     ("Demographics", "Special Purpose", "One record per subject", "STUDYID,USUBJID"),
    "DS":     ("Disposition", "Events", "One record per disposition event per subject",
               "STUDYID,USUBJID,DSSEQ"),
    "EX":     ("Exposure", "Interventions", "One record per dosing period per subject",
               "STUDYID,USUBJID,EXSEQ"),
    "AE":     ("Adverse Events", "Events", "One record per adverse event per subject",
               "STUDYID,USUBJID,AESEQ"),
    "CM":     ("Concomitant Medications", "Interventions",
               "One record per medication per subject", "STUDYID,USUBJID,CMSEQ"),
    "VS":     ("Vital Signs", "Findings", "One record per test per visit per subject",
               "STUDYID,USUBJID,VSSEQ"),
    "LB":     ("Laboratory Test Results", "Findings",
               "One record per test per visit per subject", "STUDYID,USUBJID,LBSEQ"),
    "SUPPAE": ("Supplemental Qualifiers for AE", "Relationship",
               "One record per supplemental qualifier per parent record",
               "STUDYID,RDOMAIN,USUBJID,IDVAR,IDVARVAL,QNAM"),
}

# Origin values in the spec -> Define-XML Origin types.
ORIGIN_MAP = {
    "Collected": "CRF",
    "Assigned": "Assigned",
    "Derived": "Derived",
    "Protocol": "Protocol",
}

# Codelists we actually apply, taken from the study's controlled terminology.
CODELISTS = {
    "SEX":    ("Sex", "text", [("M", "Male"), ("F", "Female"), ("U", "Unknown")]),
    "NY":     ("No Yes Response", "text", [("N", "No"), ("Y", "Yes")]),
    "AESEV":  ("Severity/Intensity Scale for Adverse Events", "text",
               [("MILD", "Mild"), ("MODERATE", "Moderate"), ("SEVERE", "Severe")]),
    "AEOUT":  ("Outcome of Event", "text",
               [("RECOVERED/RESOLVED", "Recovered/Resolved"),
                ("RECOVERING/RESOLVING", "Recovering/Resolving"),
                ("NOT RECOVERED/NOT RESOLVED", "Not Recovered/Not Resolved"),
                ("FATAL", "Fatal"), ("UNKNOWN", "Unknown")]),
    "NCOMPLT": ("Completion/Reason for Non-Completion", "text",
                [("COMPLETED", "Completed"), ("ADVERSE EVENT", "Adverse Event")]),
}
#  Real derivation text. The spec's "Source Dataset" column names WHERE a value
#  came from, not HOW it was produced - and "how" is exactly what a reviewer opens
#  a MethodDef to find out. These cover the derivations that get asked about;
#  anything else falls back to a plain statement of its source.
DERIVATION_TEXT = {
    "USUBJID":  "Concatenation of STUDYID, SITEID and SUBJID separated by hyphens "
                "(e.g. ABC-01-01-001). SUBJID alone is unique only within a site.",
    "AGE":      "Completed years from BRTHDTC to RFICDTC (informed consent), using "
                "continuous interval counting so a birthday not yet reached is not counted.",
    "RFSTDTC":  "Date of first study treatment: the earliest EXSTDTC for the subject. "
                "This is the reference start date from which all --DY values are derived.",
    "RFENDTC":  "Date of last study treatment: the latest EXENDTC for the subject.",
    "RFXSTDTC": "First study-treatment exposure; equal to RFSTDTC in this study.",
    "RFXENDTC": "Last study-treatment exposure; equal to RFENDTC in this study.",
    "RFICDTC":  "Informed consent date, converted from the collected DD-MMM-YYYY to ISO 8601.",
    "RFPENDTC": "End of subject participation, taken from the disposition form (EOSDT).",
    "ARMCD":    "Derived from ARM: 'Drug A' to 'A', 'Placebo' to 'P'.",
    "ACTARM":   "Actual arm. Equal to planned ARM in this study; no subject was mis-dosed.",
    "ACTARMCD": "Actual arm code. Equal to ARMCD in this study.",
    "AEDECOD":  "Dictionary-coded term for AETERM. ILLUSTRATIVE ONLY - a real value comes "
                "from MedDRA coding performed by trained coders against a licensed version.",
    "CMDECOD":  "Dictionary-coded term for CMTRT. ILLUSTRATIVE ONLY - a real value comes "
                "from WHODrug coding against a licensed version.",
    "VSSTRESN": "Standardised numeric result. Equal to the collected value in this study; "
                "no unit conversion was required.",
    "VSSTRESC": "Standardised result as character. Copies VSORRES verbatim when no unit "
                "conversion is required, preserving the precision the site recorded.",
    "LBNRIND":  "Reference range indicator, derived by comparing LBSTRESN with LBSTNRLO and "
                "LBSTNRHI: LOW if below, HIGH if above, otherwise NORMAL. Indicates the range "
                "comparison only and does not imply clinical significance.",
    "DSDECOD":  "Standardised disposition term derived from the collected end-of-study "
                "status and reason.",
    "DSCAT":    "PROTOCOL MILESTONE for informed consent and randomisation records; "
                "DISPOSITION EVENT for the record describing how the subject left the study.",
}

def derivation_text(var, source):
    """How the value was produced. Falls back to naming the source."""
    if var in DERIVATION_TEXT:
        return DERIVATION_TEXT[var]
    if var.endswith("SEQ"):
        return ("Sequence number assigned per subject after sorting the records into a "
                "deterministic order, so that USUBJID plus this variable is unique.")
    if var.endswith("DY"):
        return ("Study day relative to RFSTDTC. On or after the reference date the value is "
                "the date difference plus one; before it, the plain difference. There is no Day 0.")
    if var.endswith("BLFL"):
        return ("Baseline flag. 'Y' on the last record on or before first dose, per subject "
                "per test, for tests that also have a post-dose result. Otherwise null.")
    if var.endswith("DTC"):
        return f"Collected date from {source}, converted to ISO 8601 (YYYY-MM-DD)."
    if var.endswith("STRESU") or var.endswith("STNRLO") or var.endswith("STNRHI"):
        return f"Standardised from the collected value in {source}."
    return f"Derived from {source}."

# variable -> codelist OID
VAR_CODELIST = {
    "SEX": "SEX", "AESER": "NY", "AESEV": "AESEV", "AEOUT": "AEOUT",
    "DSDECOD": "NCOMPLT",
}


def actual_columns(dom):
    """The variables the built dataset ACTUALLY contains."""
    import csv
    path = os.path.join(HERE, "sdtm", f"{dom.lower()}.csv")
    if not os.path.exists(path):
        return None
    with open(path, newline="") as f:
        return list(next(csv.reader(f)))


def read_spec():
    """Pull variable-level metadata out of the mapping specification workbook.

    IMPORTANT: define.xml must describe what was SUBMITTED. The specification
    also documents fields that are COLLECTED BUT NOT SUBMITTED - EOSOTH, VSND
    and EXINTP - written as "(EOSOTH)" with parentheses and labelled "NOT an
    SDTM variable". Those are documentation, not variables: they exist in no
    dataset, and declaring them would tell a reviewer to expect columns that
    are not there. (They also produced Name="(EOSOTH)", which is not even a
    legal variable name.)

    So rows are kept only if the variable is genuinely present in the built
    dataset. The data is the authority; the spec supplies the metadata for it.
    """
    wb = openpyxl.load_workbook(os.path.join(HERE, "SDTM_Mapping_Specification.xlsx"))
    domains = {}
    for dom in DOMAIN_META:
        if dom not in wb.sheetnames:
            continue
        ws = wb[dom]
        cols = actual_columns(dom)          # None if the dataset is absent
        rows, header_seen, skipped = [], False, []
        for r in ws.iter_rows(values_only=True):
            vals = list(r)
            if not any(v is not None for v in vals):
                continue
            if not header_seen:
                if vals[0] == "#":            # the header row
                    header_seen = True
                continue
            # data row: #, Variable, Label, Type, Core, Origin, Source
            if vals[1] is None:
                continue
            vname = str(vals[1]).strip()
            if cols is not None and vname not in cols:
                skipped.append(vname)
                continue
            rows.append({
                "name": vname,
                "label": str(vals[2] or "").strip(),
                "type": str(vals[3] or "Char").strip(),
                "core": str(vals[4] or "").strip(),
                "origin": str(vals[5] or "").strip(),
                "source": str(vals[6] or "").strip(),
            })
        domains[dom] = rows
        if skipped:
            print(f"  {dom}: excluded {len(skipped)} non-submitted field(s): "
                  f"{', '.join(skipped)}")
    return domains


def actual_lengths(dom):
    """Read the built dataset to get REAL character lengths.

    Define-XML must state a Length for character variables. Taking it from the
    data rather than guessing means the define describes what was actually
    produced — which is the thing a reviewer will open.
    """
    import csv
    path = os.path.join(HERE, "sdtm", f"{dom.lower()}.csv")
    if not os.path.exists(path):
        return {}
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return {}
    out = {}
    for col in rows[0]:
        out[col] = max((len(r[col] or "") for r in rows), default=1) or 1
    return out


def build():
    spec = read_spec()

    root = ET.Element(q(ODM_NS, "ODM"), {
        q(XSI_NS, "schemaLocation"): f"{ODM_NS} define2-0-0.xsd",
        q(DEF_NS, "Context"): "Submission",
        "FileOID": f"{STUDYID}-define",
        "FileType": "Snapshot",
        "ODMVersion": "1.3.2",
        "CreationDateTime": "2026-07-20T00:00:00",
        "Originator": "Clinical Programming Bootcamp (synthetic training data)",
    })

    study = ET.SubElement(root, q(ODM_NS, "Study"), {"OID": f"STD.{STUDYID}"})
    gv = ET.SubElement(study, q(ODM_NS, "GlobalVariables"))
    ET.SubElement(gv, q(ODM_NS, "StudyName")).text = STUDYID
    ET.SubElement(gv, q(ODM_NS, "StudyDescription")).text = STUDYDESC
    ET.SubElement(gv, q(ODM_NS, "ProtocolName")).text = PROTOCOL

    mdv = ET.SubElement(study, q(ODM_NS, "MetaDataVersion"), {
        "OID": "MDV.1", "Name": f"{STUDYID} SDTM {STANDARD_VERSION}",
        "Description": f"{STANDARD} {STANDARD_VERSION}",
        q(DEF_NS, "DefineVersion"): DEFINE_VERSION,
        q(DEF_NS, "StandardName"): STANDARD,
        q(DEF_NS, "StandardVersion"): STANDARD_VERSION,
    })

    # --- the annotated CRF, referenced as a leaf ------------------------
    acrf = ET.SubElement(mdv, q(DEF_NS, "AnnotatedCRF"))
    dref = ET.SubElement(acrf, q(DEF_NS, "DocumentRef"), {"leafID": "LF.acrf"})
    leaf = ET.SubElement(mdv, q(DEF_NS, "leaf"),
                         {"ID": "LF.acrf", q(XLINK_NS, "href"): "ABC-01_Sample_CRF.pdf"})
    ET.SubElement(leaf, q(DEF_NS, "title")).text = "Annotated Case Report Form"

    methods = []
    #  MethodOID lives on the ItemRef, NOT the ItemDef. Build the lookup first
    #  so each ItemRef can point at the derivation that produced it - otherwise
    #  the MethodDefs sit in the file orphaned and a reviewer cannot follow them.
    derived = {}
    for dom_, rows_ in spec.items():
        for v_ in rows_:
            if ORIGIN_MAP.get(v_["origin"]) == "Derived" and v_["source"]:
                derived[(dom_, v_["name"])] = (f"MT.{dom_}.{v_['name']}", v_["source"])

    # --- one ItemGroupDef per dataset -----------------------------------
    for order, (dom, rows) in enumerate(sorted(spec.items()), start=1):
        label, cls, structure, keys = DOMAIN_META[dom]
        lens = actual_lengths(dom)
        igd = ET.SubElement(mdv, q(ODM_NS, "ItemGroupDef"), {
            "OID": f"IG.{dom}", "Name": dom, "Repeating": "No" if dom == "DM" else "Yes",
            "IsReferenceData": "No", "SASDatasetName": dom, "Domain": dom,
            "Purpose": "Tabulation",
            q(DEF_NS, "Structure"): structure,
            q(DEF_NS, "Class"): cls,
            q(DEF_NS, "ArchiveLocationID"): f"LF.{dom}",
        })
        ET.SubElement(igd, q(ODM_NS, "Description")).append(
            tr(f"{label}"))
        for i, v in enumerate(rows, start=1):
            ref = {
                "ItemOID": f"IT.{dom}.{v['name']}",
                "OrderNumber": str(i),
                "Mandatory": "Yes" if v["core"] == "Req" else "No",
            }
            if v["name"] in keys.split(","):
                ref["KeySequence"] = str(keys.split(",").index(v["name"]) + 1)
            if (dom, v["name"]) in derived:
                ref[q(DEF_NS, "MethodOID")] = derived[(dom, v["name"])][0]
            ET.SubElement(igd, q(ODM_NS, "ItemRef"), ref)
        # the dataset file itself, as a leaf
        lf = ET.SubElement(mdv, q(DEF_NS, "leaf"),
                           {"ID": f"LF.{dom}", q(XLINK_NS, "href"): f"{dom.lower()}.xpt"})
        ET.SubElement(lf, q(DEF_NS, "title")).text = f"{dom.lower()}.xpt"

    # --- one ItemDef per variable ---------------------------------------
    for dom, rows in sorted(spec.items()):
        lens = actual_lengths(dom)
        for v in rows:
            is_char = v["type"].lower().startswith("char")
            attrs = {
                "OID": f"IT.{dom}.{v['name']}", "Name": v["name"],
                "SASFieldName": v["name"],
                "DataType": "text" if is_char else "integer",
            }
            if is_char:
                attrs["Length"] = str(max(lens.get(v["name"], 1), 1))
            itd = ET.SubElement(mdv, q(ODM_NS, "ItemDef"), attrs)
            ET.SubElement(itd, q(ODM_NS, "Description")).append(tr(v["label"]))

            cl = VAR_CODELIST.get(v["name"])
            if cl:
                ET.SubElement(itd, q(ODM_NS, "CodeListRef"), {"CodeListOID": f"CL.{cl}"})

            otype = ORIGIN_MAP.get(v["origin"], "Derived")
            ET.SubElement(itd, q(DEF_NS, "Origin"), {"Type": otype})

    # --- codelists -------------------------------------------------------
    for oid, (name, dtype, items) in CODELISTS.items():
        cl = ET.SubElement(mdv, q(ODM_NS, "CodeList"), {
            "OID": f"CL.{oid}", "Name": name, "DataType": dtype})
        for code, decode in items:
            ci = ET.SubElement(cl, q(ODM_NS, "CodeListItem"), {"CodedValue": code})
            d = ET.SubElement(ci, q(ODM_NS, "Decode"))
            d.append(tr(decode))

    # --- methods (how derived variables were produced) -------------------
    for (dom, vname), (mid, how) in sorted(derived.items()):
        md = ET.SubElement(mdv, q(ODM_NS, "MethodDef"),
                           {"OID": mid, "Name": f"Derivation of {vname}", "Type": "Computation"})
        md.append(desc(derivation_text(vname, how)))

    return root


def tr(text):
    e = ET.Element(q(ODM_NS, "TranslatedText"), {"{http://www.w3.org/XML/1998/namespace}lang": "en"})
    e.text = text
    return e


def desc(text):
    d = ET.Element(q(ODM_NS, "Description"))
    d.append(tr(text))
    return d


def main():
    root = build()
    raw = ET.tostring(root, encoding="utf-8")
    pretty = minidom.parseString(raw).toprettyxml(indent="  ", encoding="UTF-8").decode("utf-8")
    # add the stylesheet PI so the file renders in a browser
    pretty = pretty.replace(
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<?xml-stylesheet type="text/xsl" href="define2-0-0.xsl"?>', 1)
    pretty = "\n".join(l for l in pretty.split("\n") if l.strip())

    path = os.path.join(OUT, "define.xml")
    with open(path, "w", encoding="utf-8") as f:
        f.write(pretty + "\n")
    print(f"wrote define/define.xml  ({len(pretty.splitlines())} lines)")

    # ---- self-check: it must parse, and contain everything from the spec ----
    tree = ET.parse(path)
    r = tree.getroot()
    igds = r.findall(f".//{q(ODM_NS, 'ItemGroupDef')}")
    itds = r.findall(f".//{q(ODM_NS, 'ItemDef')}")
    cls_ = r.findall(f".//{q(ODM_NS, 'CodeList')}")
    mts = r.findall(f".//{q(ODM_NS, 'MethodDef')}")
    spec = read_spec()
    want_vars = sum(len(v) for v in spec.values())
    print(f"  parses OK")
    print(f"  ItemGroupDef (datasets): {len(igds)}  (spec has {len(spec)})")
    print(f"  ItemDef (variables):     {len(itds)}  (spec has {want_vars})")
    print(f"  CodeList:                {len(cls_)}")
    print(f"  MethodDef (derivations): {len(mts)}")
    assert len(igds) == len(spec), "dataset count mismatch"
    assert len(itds) == want_vars, "variable count mismatch"

    #  Every declared variable must EXIST in its dataset, and every column in the
    #  dataset must be declared. define.xml describing columns that are not there
    #  (or omitting ones that are) is a conformance finding a reviewer will raise.
    for dom in spec:
        cols = actual_columns(dom)
        if cols is None:
            continue
        declared = [it.get("Name") for it in itds
                    if it.get("OID", "").startswith(f"IT.{dom}.")]
        assert set(declared) == set(cols), (
            f"{dom}: define.xml and the dataset disagree - "
            f"only in define {sorted(set(declared) - set(cols))}, "
            f"only in data {sorted(set(cols) - set(declared))}")
    print(f"  every ItemDef matches a real column in its dataset")

    #  Every MethodDef must be POINTED AT by an ItemRef. An orphaned method is
    #  a silent defect: the file looks complete but the derivation is unreachable.
    declared = {m.get("OID") for m in mts}
    referenced = {ir.get(q(DEF_NS, "MethodOID"))
                  for ir in r.findall(f".//{q(ODM_NS, 'ItemRef')}")
                  if ir.get(q(DEF_NS, "MethodOID"))}
    orphan = declared - referenced
    dangling = referenced - declared
    print(f"  MethodDef referenced:    {len(referenced)} of {len(declared)}")
    assert not orphan, f"MethodDefs never referenced: {sorted(orphan)[:3]}"
    assert not dangling, f"ItemRefs point at missing methods: {sorted(dangling)[:3]}"
    print("\n  NOTE: structurally correct and well-formed, but NOT schema-validated.")
    print("        Validate with Pinnacle 21 against define2-0-0.xsd before any real use.")


def render_html():
    """Also write a PRE-RENDERED define.html.

    Chromium browsers (Chrome, Edge) refuse to apply an XSLT stylesheet over
    file:// - every local file is treated as its own origin, so loading the .xsl
    is a cross-origin request and is blocked. The user sees raw XML or a blank
    page. Firefox still allows it, and serving over http:// works everywhere,
    but neither is reasonable to ask of a room of trainees.

    So we transform it here and ship the HTML alongside. define.xml remains the
    real artifact; define.html is a convenience copy that opens by double-click
    in any browser.
    """
    try:
        import lxml.etree as LET
    except ImportError:
        print("  (lxml not installed - skipped define.html; "
              "install lxml or serve define.xml over http://)")
        return
    xml_path = os.path.join(OUT, "define.xml")
    xsl_path = os.path.join(OUT, "define2-0-0.xsl")
    if not os.path.exists(xsl_path):
        print("  (define2-0-0.xsl not found - skipped define.html)")
        return
    transform = LET.XSLT(LET.parse(xsl_path))
    html = str(transform(LET.parse(xml_path)))
    banner = ("<p style=\"background:#FDF0D5;border-left:4px solid #E8833A;"
              "padding:9px 12px;font-size:12.5px;margin:14px 0\">"
              "<b>This is a pre-rendered copy.</b> The real artifact is "
              "<code>define.xml</code>, which a browser styles on the fly using "
              "<code>define2-0-0.xsl</code>. Chrome and Edge block that over "
              "<code>file://</code>, so this HTML is provided for convenience.</p>")
    html = html.replace("<body>", "<body>", 1)
    html = html.replace('<div class="wrap">', '<div class="wrap">' + banner, 1)
    out = os.path.join(OUT, "define.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    rows = html.count("<tr")
    print(f"wrote define/define.html  ({rows} table rows rendered)")


if __name__ == "__main__":
    main()
    render_html()
