#!/usr/bin/env python3
"""
build_adam_define_xml.py — generate define_adam.xml (Define-XML v2.0, ADaM-IG v1.2)
for study ABC-01, from ADaM_Specification.xlsx.

The ADaM counterpart of build_define_xml.py. Same principle: the metadata lives in
the workbook; this turns it into define.xml. Change the spec, re-run, the define
follows. It adds the piece the SDTM define does not have — VALUE-LEVEL METADATA:
for the BDS/TTE datasets, AVAL means something different for each PARAMCD, so each
parameter gets its own ItemDef, WhereClause and derivation Method.

⚠️  SCOPE / HONESTY
Structurally correct, well-formed Define-XML v2.0 suitable for TEACHING. NOT
schema-validated against define2-0-0.xsd — that needs the XSD and Pinnacle 21.
The script checks that the XML parses and that every dataset, variable and
value-level entry in the spec made it in.

Run (after build_adam_spec_xlsx.py):  python3 build_adam_define_xml.py
Writes: define/define_adam.xml
"""

import csv
import os
from xml.etree import ElementTree as ET
from xml.dom import minidom

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "define")
os.makedirs(OUT, exist_ok=True)
SPEC = os.path.join(HERE, "ADaM_Specification.xlsx")

STUDYID = "ABC-01"
STUDYDESC = ("Synthetic training study. Randomised, double-blind, placebo-controlled, "
             "8 subjects at 2 sites, 4 weeks of treatment.")
PROTOCOL = "ABC-01"
DEFINE_VERSION = "2.0.0"
STANDARD = "ADaM-IG"
STANDARD_VERSION = "1.2"

ODM_NS = "http://www.cdisc.org/ns/odm/v1.3"
DEF_NS = "http://www.cdisc.org/ns/def/v2.0"
XSI_NS = "http://www.w3.org/2001/XMLSchema-instance"
XLINK_NS = "http://www.w3.org/1999/xlink"
XML_NS = "http://www.w3.org/XML/1998/namespace"
ET.register_namespace("", ODM_NS)
ET.register_namespace("def", DEF_NS)
ET.register_namespace("xsi", XSI_NS)
ET.register_namespace("xlink", XLINK_NS)


def q(ns, tag):
    return f"{{{ns}}}{tag}"


# def:Class values per the ADaM Define-XML controlled terminology.
CLASS_CT = {
    "ADSL": "SUBJECT LEVEL ANALYSIS DATASET",
    "ADAE": "OCCURRENCE DATA STRUCTURE",
    "ADVS": "BASIC DATA STRUCTURE",
    "ADLB": "BASIC DATA STRUCTURE",
    "ADTTE": "BASIC DATA STRUCTURE",
}
# DataType -> whether it needs a Length (character only).
CHAR_TYPES = {"text"}


def read_spec():
    """Read the workbook: dataset meta, per-variable rows, value-level, codelists."""
    wb = openpyxl.load_workbook(SPEC, read_only=True)

    # Index sheet: Dataset, Label, Class, Structure, Key
    datasets = {}
    ws = wb["Index"]
    header_seen = False
    for r in ws.iter_rows(values_only=True):
        v = list(r)
        if not header_seen:
            if v and v[0] == "Dataset":
                header_seen = True
            continue
        if not v or v[0] is None:
            continue
        datasets[str(v[0]).strip()] = {
            "label": str(v[1] or "").strip(),
            "structure": str(v[3] or "").strip(),
            "key": [k.strip() for k in str(v[4] or "").split(",") if k.strip()],
            "vars": [],
        }

    # per-dataset variable sheets
    for ds in datasets:
        ws = wb[ds]
        header_seen = False
        for r in ws.iter_rows(values_only=True):
            v = list(r)
            if not header_seen:
                if v and v[0] == "#":
                    header_seen = True
                continue
            if not v or v[0] is None:
                continue
            datasets[ds]["vars"].append({
                "name": str(v[1]).strip(),
                "label": str(v[2] or "").strip(),
                "type": str(v[3] or "text").strip(),
                "origin": str(v[4] or "").strip(),
                "codelist": str(v[5] or "").strip(),
                "deriv": str(v[6] or "").strip(),
            })

    # value-level: Dataset, PARAMCD, PARAM, PARAMN, Unit, AVAL Derivation
    vlm = {}
    ws = wb["ValueLevel"]
    header_seen = False
    for r in ws.iter_rows(values_only=True):
        v = list(r)
        if not header_seen:
            if v and v[0] == "Dataset":
                header_seen = True
            continue
        if not v or v[0] is None:
            continue
        vlm.setdefault(str(v[0]).strip(), []).append({
            "paramcd": str(v[1]).strip(), "param": str(v[2]).strip(),
            "paramn": v[3], "unit": str(v[4] or "").strip(),
            "deriv": str(v[5] or "").strip(),
        })

    # codelists: Codelist, Name, Coded Value, Decode  (Codelist/Name only on first row)
    codelists = {}
    ws = wb["Codelists"]
    header_seen, cur = False, None
    for r in ws.iter_rows(values_only=True):
        v = list(r)
        if not header_seen:
            if v and v[0] == "Codelist":
                header_seen = True
            continue
        if not v or (v[2] is None):
            continue
        if v[0]:
            cur = str(v[0]).strip()
            codelists[cur] = {"name": str(v[1] or "").strip(), "items": []}
        codelists[cur]["items"].append((str(v[2]).strip(), str(v[3] or "").strip()))

    wb.close()
    return datasets, vlm, codelists


def actual_lengths(ds):
    path = os.path.join(HERE, "adam", f"{ds.lower()}.csv")
    if not os.path.exists(path):
        return {}
    with open(path, newline="") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return {}
    return {c: max((len(r[c] or "") for r in rows), default=1) or 1 for c in rows[0]}


def tr(text):
    e = ET.Element(q(ODM_NS, "TranslatedText"), {q(XML_NS, "lang"): "en"})
    e.text = text
    return e


def desc(text):
    d = ET.Element(q(ODM_NS, "Description"))
    d.append(tr(text))
    return d


def datatype_attr(t):
    return t if t in ("text", "integer", "float", "date") else "text"


def build():
    datasets, vlm, codelists = read_spec()

    root = ET.Element(q(ODM_NS, "ODM"), {
        q(XSI_NS, "schemaLocation"): f"{ODM_NS} define2-0-0.xsd",
        q(DEF_NS, "Context"): "Submission",
        "FileOID": f"{STUDYID}-define-adam",
        "FileType": "Snapshot",
        "ODMVersion": "1.3.2",
        "CreationDateTime": "2026-07-21T00:00:00",
        "Originator": "Clinical Programming Bootcamp (synthetic training data)",
    })
    study = ET.SubElement(root, q(ODM_NS, "Study"), {"OID": f"STD.{STUDYID}"})
    gv = ET.SubElement(study, q(ODM_NS, "GlobalVariables"))
    ET.SubElement(gv, q(ODM_NS, "StudyName")).text = STUDYID
    ET.SubElement(gv, q(ODM_NS, "StudyDescription")).text = STUDYDESC
    ET.SubElement(gv, q(ODM_NS, "ProtocolName")).text = PROTOCOL

    mdv = ET.SubElement(study, q(ODM_NS, "MetaDataVersion"), {
        "OID": "MDV.1", "Name": f"{STUDYID} ADaM {STANDARD_VERSION}",
        "Description": f"{STANDARD} {STANDARD_VERSION}",
        q(DEF_NS, "DefineVersion"): DEFINE_VERSION,
        q(DEF_NS, "StandardName"): STANDARD,
        q(DEF_NS, "StandardVersion"): STANDARD_VERSION,
    })

    methods = []          # (oid, name, text)  collected, emitted at the end
    valuelists = []       # (oid, [(item_oid, wc_oid, mandatory)])
    wheres = []           # (oid, itemoid_paramcd, value)
    vlm_itemdefs = []     # (oid, name, dtype, length, label, method_oid, codelist)

    # --- ItemGroupDef per dataset ---------------------------------------
    for order, ds in enumerate(datasets, start=1):
        meta = datasets[ds]
        igd = ET.SubElement(mdv, q(ODM_NS, "ItemGroupDef"), {
            "OID": f"IG.{ds}", "Name": ds,
            "Repeating": "No" if ds == "ADSL" else "Yes",
            "IsReferenceData": "No", "SASDatasetName": ds,
            "Purpose": "Analysis",
            q(DEF_NS, "Structure"): meta["structure"],
            q(DEF_NS, "Class"): CLASS_CT[ds],
            q(DEF_NS, "ArchiveLocationID"): f"LF.{ds}",
        })
        igd.append(desc(meta["label"]))
        for i, v in enumerate(meta["vars"], start=1):
            ref = {"ItemOID": f"IT.{ds}.{v['name']}", "OrderNumber": str(i),
                   "Mandatory": "Yes" if v["name"] in meta["key"] else "No"}
            if v["name"] in meta["key"]:
                ref["KeySequence"] = str(meta["key"].index(v["name"]) + 1)
            if v["origin"] in ("Derived", "Assigned") and v["deriv"]:
                mid = f"MT.{ds}.{v['name']}"
                ref[q(DEF_NS, "MethodOID")] = mid
                methods.append((mid, f"Algorithm for {v['name']}", v["deriv"]))
            ET.SubElement(igd, q(ODM_NS, "ItemRef"), ref)
        lf = ET.SubElement(mdv, q(DEF_NS, "leaf"),
                           {"ID": f"LF.{ds}", q(XLINK_NS, "href"): f"{ds.lower()}.xpt"})
        ET.SubElement(lf, q(DEF_NS, "title")).text = f"{ds.lower()}.xpt"

    # --- ItemDef per variable -------------------------------------------
    for ds in datasets:
        lens = actual_lengths(ds)
        for v in datasets[ds]["vars"]:
            dtype = datatype_attr(v["type"])
            attrs = {"OID": f"IT.{ds}.{v['name']}", "Name": v["name"],
                     "SASFieldName": v["name"], "DataType": dtype}
            if dtype in CHAR_TYPES:
                attrs["Length"] = str(max(lens.get(v["name"], 1), 1))
            itd = ET.SubElement(mdv, q(ODM_NS, "ItemDef"), attrs)
            itd.append(desc(v["label"]))
            if v["codelist"]:
                ET.SubElement(itd, q(ODM_NS, "CodeListRef"),
                              {"CodeListOID": f"CL.{v['codelist']}"})
            # AVAL in a value-level dataset points at its ValueListDef.
            if v["name"] == "AVAL" and ds in vlm:
                ET.SubElement(itd, q(DEF_NS, "ValueListRef"),
                              {"ValueListOID": f"VL.{ds}.AVAL"})
            ET.SubElement(itd, q(DEF_NS, "Origin"), {"Type": v["origin"] or "Derived"})

    # --- value-level metadata: one AVAL ItemDef per PARAMCD -------------
    for ds, items in vlm.items():
        vl_items = []
        for it in items:
            pc = it["paramcd"]
            item_oid = f"IT.{ds}.AVAL.{pc}"
            wc_oid = f"WC.{ds}.{pc}"
            mid = f"MT.{ds}.AVAL.{pc}"
            # per-parameter AVAL ItemDef (numeric)
            vlm_itemdefs.append((item_oid, "AVAL", "float", None,
                                 f"{it['param']} — analysis value", mid, ""))
            methods.append((mid, f"Algorithm for AVAL where PARAMCD={pc}", it["deriv"]))
            wheres.append((wc_oid, f"IT.{ds}.PARAMCD", pc))
            vl_items.append((item_oid, wc_oid))
        valuelists.append((f"VL.{ds}.AVAL", vl_items))

    for oid, name, dtype, length, label, mid, cl in vlm_itemdefs:
        attrs = {"OID": oid, "Name": name, "SASFieldName": name,
                 "DataType": datatype_attr(dtype)}
        itd = ET.SubElement(mdv, q(ODM_NS, "ItemDef"), attrs)
        itd.append(desc(label))
        ET.SubElement(itd, q(DEF_NS, "Origin"), {"Type": "Derived"})

    # --- ValueListDefs ---------------------------------------------------
    for vl_oid, vl_items in valuelists:
        vld = ET.SubElement(mdv, q(DEF_NS, "ValueListDef"), {"OID": vl_oid})
        for i, (item_oid, wc_oid) in enumerate(vl_items, start=1):
            ref = ET.SubElement(vld, q(ODM_NS, "ItemRef"),
                                {"ItemOID": item_oid, "OrderNumber": str(i),
                                 "Mandatory": "No"})
            ET.SubElement(ref, q(DEF_NS, "WhereClauseRef"), {"WhereClauseOID": wc_oid})

    # --- WhereClauseDefs -------------------------------------------------
    for wc_oid, item_oid, value in wheres:
        wcd = ET.SubElement(mdv, q(DEF_NS, "WhereClauseDef"), {"OID": wc_oid})
        rc = ET.SubElement(wcd, q(ODM_NS, "RangeCheck"),
                           {"Comparator": "EQ", q(DEF_NS, "SoftHard"): "Soft",
                            "ItemOID": item_oid})
        ET.SubElement(rc, q(ODM_NS, "CheckValue")).text = value

    # --- CodeLists -------------------------------------------------------
    for oid, cl in codelists.items():
        # numeric-looking codelists still carry text coded values here for simplicity
        cle = ET.SubElement(mdv, q(ODM_NS, "CodeList"),
                            {"OID": f"CL.{oid}", "Name": cl["name"], "DataType": "text"})
        for code, decode in cl["items"]:
            ci = ET.SubElement(cle, q(ODM_NS, "CodeListItem"), {"CodedValue": code})
            d = ET.SubElement(ci, q(ODM_NS, "Decode"))
            d.append(tr(decode))

    # --- MethodDefs (dedup by OID) --------------------------------------
    seen = set()
    for mid, name, text in methods:
        if mid in seen:
            continue
        seen.add(mid)
        md = ET.SubElement(mdv, q(ODM_NS, "MethodDef"),
                           {"OID": mid, "Name": name, "Type": "Computation"})
        md.append(desc(text))

    return root, datasets, vlm


def main():
    root, datasets, vlm = build()
    raw = ET.tostring(root, encoding="utf-8")
    pretty = minidom.parseString(raw).toprettyxml(indent="  ", encoding="UTF-8").decode("utf-8")
    pretty = pretty.replace(
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<?xml-stylesheet type="text/xsl" href="define2-0-0.xsl"?>', 1)
    pretty = "\n".join(l for l in pretty.split("\n") if l.strip())

    path = os.path.join(OUT, "define_adam.xml")
    with open(path, "w", encoding="utf-8") as f:
        f.write(pretty + "\n")
    print(f"wrote define/define_adam.xml  ({len(pretty.splitlines())} lines)")

    # ---- self-check ----
    tree = ET.parse(path)
    r = tree.getroot()
    igds = r.findall(f".//{q(ODM_NS, 'ItemGroupDef')}")
    itds = r.findall(f".//{q(ODM_NS, 'ItemDef')}")
    cls_ = r.findall(f".//{q(ODM_NS, 'CodeList')}")
    mts = r.findall(f".//{q(ODM_NS, 'MethodDef')}")
    vls = r.findall(f".//{q(DEF_NS, 'ValueListDef')}")
    wcs = r.findall(f".//{q(DEF_NS, 'WhereClauseDef')}")
    want_vars = sum(len(d["vars"]) for d in datasets.values())
    want_vlm = sum(len(v) for v in vlm.values())
    print("  parses OK")
    print(f"  ItemGroupDef (datasets): {len(igds)}  (spec has {len(datasets)})")
    print(f"  ItemDef (variables):     {len(itds)}  (spec {want_vars} vars + {want_vlm} value-level)")
    print(f"  ValueListDef:            {len(vls)}")
    print(f"  WhereClauseDef:          {len(wcs)}  (one per parameter)")
    print(f"  CodeList:                {len(cls_)}")
    print(f"  MethodDef:               {len(mts)}")
    assert len(igds) == len(datasets), "dataset count mismatch"
    assert len(itds) == want_vars + want_vlm, "ItemDef count mismatch"
    assert len(wcs) == want_vlm, "WhereClause count mismatch"
    # every ValueListRef resolves to a ValueListDef, every WhereClauseRef to a def
    vl_oids = {v.get("OID") for v in vls}
    for ref in r.findall(f".//{q(DEF_NS, 'ValueListRef')}"):
        assert ref.get("ValueListOID") in vl_oids, f"dangling ValueListRef {ref.get('ValueListOID')}"
    wc_oids = {w.get("OID") for w in wcs}
    for ref in r.findall(f".//{q(DEF_NS, 'WhereClauseRef')}"):
        assert ref.get("WhereClauseOID") in wc_oids, "dangling WhereClauseRef"
    print("  self-check: all value-list and where-clause references resolve")


if __name__ == "__main__":
    main()
