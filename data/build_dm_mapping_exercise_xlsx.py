#!/usr/bin/env python3
"""
build_dm_mapping_exercise_xlsx.py — hands-on Excel exercise: map raw -> SDTM DM by hand.

Trainees complete the DM domain in a spreadsheet BEFORE writing any code, so the
programming step automates something they already understand.

Sheets
  Instructions        what to do, the rules, the legend
  RAW dm / ex / ds    the three source forms (read-only)
  CT Lookups          the controlled-terminology maps they need
  SDTM DM (EXERCISE)  subject 1 worked as an example; 7 rows to complete (yellow)
  Check               LIVE formulas: per-cell OK/? and a % complete
  ANSWER              the finished domain (instructor copy — delete for trainees)

Run: python3 build_dm_mapping_exercise_xlsx.py
"""

import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "DM_Mapping_Exercise.xlsx")

FONT = "Arial"
INK, TEAL, ACCENT = "0F2E3D", "0E7C86", "E8833A"
PAPER, LINE = "F3F7F8", "CFDEE1"
YELLOW = "FFF2CC"          # cells the trainee fills in
GREEN = "E2EFDA"           # worked example
WARN = "FDF1E7"
H_FILL = PatternFill("solid", fgColor=INK)
Y_FILL = PatternFill("solid", fgColor=YELLOW)
G_FILL = PatternFill("solid", fgColor=GREEN)
BAND = PatternFill("solid", fgColor=PAPER)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read(name):
    with open(os.path.join(HERE, name)) as f:
        return list(csv.reader(f))


def sheet_from_csv(wb, title, name, note):
    ws = wb.create_sheet(title)
    rows = read(name)
    ws["A1"] = f"{title}  —  source file: {name}"
    ws["A1"].font = Font(name=FONT, bold=True, size=13, color=INK)
    ws["A2"] = note
    ws["A2"].font = Font(name=FONT, italic=True, size=9.5, color="5A7682")
    for j, h in enumerate(rows[0], start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
        c.fill = H_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows[1:], start=5):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
            if i % 2 == 0:
                c.fill = BAND
    for j, h in enumerate(rows[0], start=1):
        width = max(len(h), *(len(r[j - 1]) for r in rows[1:])) + 3
        ws.column_dimensions[get_column_letter(j)].width = min(width, 30)
    ws.freeze_panes = "A5"
    return ws


wb = Workbook()

# ============================================================ INSTRUCTIONS
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
ws["A1"] = "Exercise — map raw data into SDTM DM by hand"
ws["A1"].font = Font(name=FONT, bold=True, size=18, color=INK)
ws["A2"] = "Study ABC-01 · do this BEFORE writing any code"
ws["A2"].font = Font(name=FONT, size=12, color=TEAL)

blocks = [
    ("What you are doing",
     "The 'SDTM DM (EXERCISE)' sheet has the 23 SDTM DM variables as columns and one row per subject. "
     "Subject ABC-01-01-001 is already filled in as a worked example. Complete the other 7 rows by hand."),
    ("Why by hand first",
     "A spreadsheet forces you to confront every value individually. Once you have mapped 8 subjects "
     "manually, the SAS program in Notebook 04 is just automating something you already understand."),
    ("Where the data comes from",
     "THREE source sheets: 'RAW dm' (demographics), 'RAW ex' (exposure — the dose dates), and "
     "'RAW ds' (disposition — end of study). DM is not built from demographics alone."),
    ("How to check yourself",
     "The 'Check' sheet compares every cell you enter against the correct answer and shows OK or '?'. "
     "It updates as you type. Aim for 100%."),
]
r = 4
for title, body in blocks:
    ws.cell(row=r, column=1, value=title).font = Font(name=FONT, bold=True, size=11, color=TEAL)
    c = ws.cell(row=r + 1, column=1, value=body)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 2, end_column=8)
    r += 4

# legend
ws.cell(row=r, column=1, value="LEGEND").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
for fill, label in [(Y_FILL, "Yellow  =  YOU fill this in"),
                    (G_FILL, "Green   =  worked example, already done"),
                    (PatternFill("solid", fgColor=PAPER), "Grey    =  read-only source data")]:
    c = ws.cell(row=r, column=1, value="")
    c.fill = fill
    c.border = BORDER
    ws.cell(row=r, column=2, value=label).font = Font(name=FONT, size=10)
    r += 1

r += 1
ws.cell(row=r, column=1, value="THE RULES YOU NEED").font = Font(name=FONT, bold=True, size=11, color=INK)
r += 1
rules = [
    ("STUDYID / DOMAIN", 'Constant "ABC-01" and "DM".'),
    ("USUBJID", 'STUDYID + "-" + SITEID + "-" + SUBJID   ->   ABC-01-01-001. Keep the leading zeros!'),
    ("SUBJID / SITEID / COUNTRY / ARM", "Copy from RAW dm exactly as collected."),
    ("AGE", "COMPLETED years from BRTHDTC to RFICDTC. Both are DD-MMM-YYYY (e.g. 14-MAY-1969). "
            "Subtract 1 if the birthday has not happened yet by the consent date. AGEU is always YEARS."),
    ("SEX / RACE / ETHNIC", "Look up in the 'CT Lookups' sheet. Trim spaces, match the CT spelling exactly."),
    ("ARMCD", "Drug A -> A, Placebo -> P. ACTARM = ARM and ACTARMCD = ARMCD in this study."),
    ("RFSTDTC / RFENDTC", "From 'RAW ex': first and last dose date. RFXSTDTC = RFSTDTC, RFXENDTC = RFENDTC."),
    ("RFICDTC", "From 'RAW dm' (consent date)."),
    ("RFPENDTC", "From 'RAW ds' (EOSDT) — the DISPOSITION form, not demographics."),
    ("DTHDTC / DTHFL", "Leave blank — nobody died in this study."),
    ("All dates", "The RAW dates are DD-MMM-YYYY (e.g. 01-MAR-2024). SDTM wants ISO 8601: YYYY-MM-DD. CONVERT EVERY ONE — that is a real part of the job."),
]
for k, v in rules:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=9.5, color=TEAL)
    c = ws.cell(row=r, column=2, value=v)
    c.font = Font(name=FONT, size=9.5)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = 26
    r += 1

r += 1
c = ws.cell(row=r, column=1, value="INSTRUCTOR NOTE: the 'ANSWER' sheet holds the completed domain. "
                                   "Delete it before giving this file to trainees if you want them to work blind "
                                   "(the Check sheet will then stop working).")
c.font = Font(name=FONT, bold=True, italic=True, size=9.5, color="B5651A")
c.fill = PatternFill("solid", fgColor=WARN)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=8)

ws.column_dimensions["A"].width = 34
for col in "BCDEFGH":
    ws.column_dimensions[col].width = 14

# ============================================================ RAW SHEETS
sheet_from_csv(wb, "RAW dm", "dm_raw.csv",
               "Demographics form, completed at SCREENING. SEX is a code, RACE is free text, and the dates are DD-MMM-YYYY — NOT ISO.")
sheet_from_csv(wb, "RAW ex", "ex_raw.csv",
               "Exposure form. EXSTDTC / EXENDTC give you RFSTDTC / RFENDTC — convert them to ISO.")
sheet_from_csv(wb, "RAW ds", "ds_raw.csv",
               "Disposition form, completed at END OF STUDY. EOSDT gives you RFPENDTC — convert it to ISO.")

# ============================================================ CT LOOKUPS
ws = wb.create_sheet("CT Lookups")
ws["A1"] = "Controlled Terminology lookups"
ws["A1"].font = Font(name=FONT, bold=True, size=13, color=INK)
ws["A2"] = "Match the CT value EXACTLY — spelling and case."
ws["A2"].font = Font(name=FONT, italic=True, size=9.5, color="5A7682")
ct = [("SEX", "1", "M"), ("SEX", "2", "F"),
      ("RACE", "White", "WHITE"), ("RACE", "Asian", "ASIAN"),
      ("RACE", "asian", "ASIAN"),
      ("RACE", "black or african american", "BLACK OR AFRICAN AMERICAN"),
      ("RACE", "Black or African American", "BLACK OR AFRICAN AMERICAN"),
      ("ETHNIC", "Not Hispanic or Latino", "NOT HISPANIC OR LATINO"),
      ("ETHNIC", "Hispanic or Latino", "HISPANIC OR LATINO"),
      ("ETHNIC", "Unknown", "UNKNOWN"),
      ("ARMCD", "Drug A", "A"), ("ARMCD", "Placebo", "P")]
for j, h in enumerate(["Variable", "Raw value", "SDTM value"], start=1):
    c = ws.cell(row=4, column=j, value=h)
    c.font = Font(name=FONT, bold=True, size=10, color="FFFFFF"); c.fill = H_FILL; c.border = BORDER
for i, row in enumerate(ct, start=5):
    for j, v in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = Font(name=FONT, size=10, bold=(j == 3)); c.border = BORDER
        if i % 2 == 0:
            c.fill = BAND
for col, w in zip("ABC", [14, 30, 32]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================================================ EXERCISE + ANSWER
dm_sdtm = read(os.path.join("sdtm", "dm.csv"))
HEADERS = dm_sdtm[0]
ANSWER_ROWS = dm_sdtm[1:]
NSUB = len(ANSWER_ROWS)

# --- ANSWER sheet (built first so Check can reference it) ---
wsa = wb.create_sheet("ANSWER")
wsa["A1"] = "ANSWER — completed SDTM DM (instructor copy)"
wsa["A1"].font = Font(name=FONT, bold=True, size=13, color="B5651A")
wsa["A2"] = "Delete this sheet to make the trainee version blind."
wsa["A2"].font = Font(name=FONT, italic=True, size=9.5, color="5A7682")
for j, h in enumerate(HEADERS, start=1):
    c = wsa.cell(row=4, column=j, value=h)
    c.font = Font(name=FONT, bold=True, size=9.5, color="FFFFFF"); c.fill = H_FILL; c.border = BORDER
for i, row in enumerate(ANSWER_ROWS, start=5):
    for j, v in enumerate(row, start=1):
        c = wsa.cell(row=i, column=j, value=v)
        c.font = Font(name=FONT, size=9.5); c.border = BORDER
for j, h in enumerate(HEADERS, start=1):
    wsa.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 12)
wsa.freeze_panes = "C5"

# --- EXERCISE sheet ---
wse = wb.create_sheet("SDTM DM (EXERCISE)", 4)      # position it after the CT sheet
wse["A1"] = "SDTM DM — complete the yellow cells"
wse["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
wse["A2"] = ("Row 5 (green) is a worked example. Fill rows 6-12 using RAW dm / RAW ex / RAW ds "
             "and the CT Lookups. The Check sheet grades you live.")
wse["A2"].font = Font(name=FONT, italic=True, size=9.5, color="5A7682")
for j, h in enumerate(HEADERS, start=1):
    c = wse.cell(row=4, column=j, value=h)
    c.font = Font(name=FONT, bold=True, size=9.5, color="FFFFFF"); c.fill = H_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal="center", wrap_text=True)
for i in range(NSUB):
    excel_row = 5 + i
    for j, h in enumerate(HEADERS, start=1):
        if i == 0:                                  # worked example, fully filled
            c = wse.cell(row=excel_row, column=j, value=ANSWER_ROWS[0][j - 1])
            c.fill = G_FILL
        elif h in ("SUBJID", "SITEID"):             # give them the row key so rows line up
            c = wse.cell(row=excel_row, column=j, value=ANSWER_ROWS[i][j - 1])
            c.fill = BAND
        else:
            c = wse.cell(row=excel_row, column=j, value=None)
            c.fill = Y_FILL
        c.font = Font(name=FONT, size=9.5)
        c.border = BORDER
for j, h in enumerate(HEADERS, start=1):
    wse.column_dimensions[get_column_letter(j)].width = max(len(h) + 2, 12)
wse.freeze_panes = "C5"
wse.cell(row=NSUB + 7, column=1,
         value="Tip: DTHDTC and DTHFL stay empty — the Check sheet expects blanks there."
         ).font = Font(name=FONT, italic=True, size=9.5, color="5A7682")

# ============================================================ CHECK (live formulas)
wsc = wb.create_sheet("Check")
wsc["A1"] = "Check your work"
wsc["A1"].font = Font(name=FONT, bold=True, size=14, color=INK)
wsc["A2"] = "Live: each cell compares your entry to the answer. OK = correct, ? = not yet right."
wsc["A2"].font = Font(name=FONT, italic=True, size=9.5, color="5A7682")

wsc["A4"] = "Cells correct"
wsc["A4"].font = Font(name=FONT, bold=True, size=11, color=TEAL)
total_cells = (NSUB - 1) * len(HEADERS)
# count of OK across the grid (grid is written below, rows 8..)
first_grid_row, last_grid_row = 8, 8 + (NSUB - 2)
grid_range = f"B{first_grid_row}:{get_column_letter(len(HEADERS) + 1)}{last_grid_row}"
wsc["B4"] = f'=COUNTIF({grid_range},"OK")'
wsc["B4"].font = Font(name=FONT, bold=True, size=11)
wsc["C4"] = f"of {total_cells}"
wsc["C4"].font = Font(name=FONT, size=10, color="5A7682")
wsc["A5"] = "Percent complete"
wsc["A5"].font = Font(name=FONT, bold=True, size=11, color=TEAL)
wsc["B5"] = f"=IFERROR(B4/{total_cells},0)"
wsc["B5"].number_format = "0.0%"
wsc["B5"].font = Font(name=FONT, bold=True, size=11)
# explain the non-zero starting score so nobody thinks the grader is broken
note = wsc.cell(row=5, column=3,
                value="You start at 28/161 (17.4%), not 0 — SUBJID and SITEID are given to you as row "
                      "keys, and DTHDTC/DTHFL are correctly blank. The other 133 cells are yours.")
note.font = Font(name=FONT, italic=True, size=9, color="5A7682")
note.alignment = Alignment(wrap_text=True, vertical="center")
wsc.merge_cells(start_row=5, start_column=3, end_row=5, end_column=14)
wsc.row_dimensions[5].height = 26

# grid header
wsc.cell(row=7, column=1, value="Subject").font = Font(name=FONT, bold=True, size=9.5, color="FFFFFF")
wsc.cell(row=7, column=1).fill = H_FILL
wsc.cell(row=7, column=1).border = BORDER
for j, h in enumerate(HEADERS, start=2):
    c = wsc.cell(row=7, column=j, value=h)
    c.font = Font(name=FONT, bold=True, size=8.5, color="FFFFFF"); c.fill = H_FILL; c.border = BORDER
    c.alignment = Alignment(textRotation=90, horizontal="center")
wsc.row_dimensions[7].height = 70

# one row per subject the trainee must complete (rows 2..8 of the exercise)
for i in range(1, NSUB):
    xl = 5 + i                       # row on the exercise/answer sheets
    out = 7 + i                      # row on the check sheet
    c = wsc.cell(row=out, column=1, value=f"={get_column_letter(3)}'SDTM DM (EXERCISE)'!A{xl}" if False
                 else f"='SDTM DM (EXERCISE)'!C{xl}")
    c.font = Font(name=FONT, size=9.5, bold=True); c.border = BORDER
    for j in range(1, len(HEADERS) + 1):
        col = get_column_letter(j)
        f = (f"=IF(EXACT(TRIM('SDTM DM (EXERCISE)'!{col}{xl}&\"\"),"
             f"TRIM(ANSWER!{col}{xl}&\"\")),\"OK\",\"?\")")
        cc = wsc.cell(row=out, column=j + 1, value=f)
        cc.font = Font(name=FONT, size=9)
        cc.border = BORDER
        cc.alignment = Alignment(horizontal="center")
wsc.column_dimensions["A"].width = 18
for j in range(2, len(HEADERS) + 2):
    wsc.column_dimensions[get_column_letter(j)].width = 5.5
wsc.freeze_panes = "B8"

# colour OK green / ? red
wsc.conditional_formatting.add(grid_range,
    CellIsRule(operator="equal", formula=['"OK"'],
               fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100")))
wsc.conditional_formatting.add(grid_range,
    CellIsRule(operator="equal", formula=['"?"'],
               fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006")))

# --- guard: the RAW sheets must match the CSVs on disk at build time ---
import re as _re
_iso = _re.compile(r"^\d{4}-\d{2}-\d{2}$")
for _sheet, _src in [("RAW dm", "dm_raw.csv"), ("RAW ex", "ex_raw.csv"), ("RAW ds", "ds_raw.csv")]:
    _ws = wb[_sheet]
    _rows = read(_src)
    for _i, _r in enumerate(_rows[1:], start=5):
        for _j, _v in enumerate(_r, start=1):
            _got = _ws.cell(row=_i, column=_j).value
            _got = "" if _got is None else str(_got)
            assert _got == _v, f"{_sheet} r{_i}c{_j}: '{_got}' != CSV '{_v}'"
            assert not _iso.match(_got), (
                f"{_sheet} r{_i}c{_j} is ISO ('{_got}') — raw data must NOT be ISO 8601")
print("guard: RAW sheets match the CSVs and contain no ISO dates")

wb.save(OUT)
print("wrote", OUT)
print("sheets:", ", ".join(wb.sheetnames))
print(f"exercise: {NSUB - 1} subjects x {len(HEADERS)} variables = {total_cells} cells to complete")
