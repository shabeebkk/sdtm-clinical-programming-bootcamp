#!/usr/bin/env python3
"""
build_adam_spec_xlsx.py — write ADaM_Specification.xlsx for study ABC-01.

The ADaM counterpart of build_mapping_spec_xlsx.py. It holds the metadata for
ADSL, ADAE, ADVS, ADLB and ADTTE as data, then writes the workbook the
programmers map from AND that build_adam_define_xml.py turns into define.xml —
one source, two consumers.

Sheets:
  Index         one row per dataset (label, class, structure, key)
  ADSL/…/ADTTE  one row per variable (name, label, type, origin, derivation, codelist)
  ValueLevel    per-PARAMCD metadata for the BDS/TTE datasets (the ADaM value-level layer)
  Codelists     the controlled terminology the datasets apply

ALL DATA IS SYNTHETIC. Dictionary-coded terms (AEDECOD) are illustrative only.

Run:    python3 build_adam_spec_xlsx.py
Writes: ADaM_Specification.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "ADaM_Specification.xlsx")

# ---------------------------------------------------------------- metadata

# Dataset-level metadata: label, class, structure, key.
DATASETS = {
    "ADSL":  ("Subject-Level Analysis Dataset", "ADSL",
              "One record per subject", "STUDYID, USUBJID"),
    "ADAE":  ("Adverse Events Analysis Dataset", "OCCDS (Occurrence)",
              "One record per adverse event per subject", "STUDYID, USUBJID, AESEQ"),
    "ADVS":  ("Vital Signs Analysis Dataset", "BDS (Basic Data Structure)",
              "One record per subject per parameter per analysis visit",
              "STUDYID, USUBJID, PARAMCD, AVISITN"),
    "ADLB":  ("Laboratory Analysis Dataset", "BDS (Basic Data Structure)",
              "One record per subject per parameter per analysis visit",
              "STUDYID, USUBJID, PARAMCD, AVISITN"),
    "ADTTE": ("Time-to-Event Analysis Dataset", "BDS (Basic Data Structure)",
              "One record per subject per parameter", "STUDYID, USUBJID, PARAMCD"),
}

# A library of variable metadata, keyed by name, so the recurring ADSL-carried
# variables are defined once. Each entry: (label, datatype, origin, derivation).
#   datatype : text | integer | float | date   (define.xml DataType)
#   origin   : Predecessor | Derived | Assigned   (ADaM Define-XML origin types)
# "Predecessor" means copied unchanged from a predecessor dataset variable; the
# derivation names it. "Derived"/"Assigned" carry a real derivation sentence.
LIB = {
    "STUDYID":  ("Study Identifier", "text", "Predecessor", "SDTM.DM.STUDYID."),
    "USUBJID":  ("Unique Subject Identifier", "text", "Predecessor", "SDTM.DM.USUBJID."),
    "SUBJID":   ("Subject Identifier for the Study", "text", "Predecessor", "SDTM.DM.SUBJID."),
    "SITEID":   ("Study Site Identifier", "text", "Predecessor", "SDTM.DM.SITEID."),
    "COUNTRY":  ("Country", "text", "Predecessor", "SDTM.DM.COUNTRY."),
    "ARM":      ("Description of Planned Arm", "text", "Predecessor", "SDTM.DM.ARM."),
    "ACTARM":   ("Description of Actual Arm", "text", "Predecessor", "SDTM.DM.ACTARM."),
    "AGE":      ("Age", "integer", "Predecessor", "SDTM.DM.AGE."),
    "AGEU":     ("Age Units", "text", "Predecessor", "SDTM.DM.AGEU."),
    "SEX":      ("Sex", "text", "Predecessor", "SDTM.DM.SEX."),
    "RACE":     ("Race", "text", "Predecessor", "SDTM.DM.RACE."),
    "ETHNIC":   ("Ethnicity", "text", "Predecessor", "SDTM.DM.ETHNIC."),

    "TRT01P":   ("Planned Treatment for Period 01", "text", "Assigned",
                 "Planned treatment, set equal to SDTM.DM.ARM."),
    "TRT01PN":  ("Planned Treatment for Period 01 (N)", "integer", "Assigned",
                 "Numeric code for TRT01P: Placebo=1, Drug A=2. Display order, not a dose."),
    "TRT01A":   ("Actual Treatment for Period 01", "text", "Assigned",
                 "Actual treatment, set equal to SDTM.DM.ACTARM."),
    "TRT01AN":  ("Actual Treatment for Period 01 (N)", "integer", "Assigned",
                 "Numeric code for TRT01A: Placebo=1, Drug A=2."),
    "TRTP":     ("Planned Treatment", "text", "Predecessor", "ADSL.TRT01P."),
    "TRTPN":    ("Planned Treatment (N)", "integer", "Predecessor", "ADSL.TRT01PN."),
    "TRTA":     ("Actual Treatment", "text", "Predecessor", "ADSL.TRT01A."),
    "TRTAN":    ("Actual Treatment (N)", "integer", "Predecessor", "ADSL.TRT01AN."),

    "AGEGR1":   ("Pooled Age Group 1", "text", "Derived",
                 "'<65' when AGE < 65, else '>=65'."),
    "AGEGR1N":  ("Pooled Age Group 1 (N)", "integer", "Derived",
                 "Numeric code for AGEGR1: '<65'=1, '>=65'=2."),

    "RFICDT":   ("Date of Informed Consent", "date", "Derived",
                 "Numeric date from SDTM.DM.RFICDTC."),
    "RANDDT":   ("Date of Randomization", "date", "Derived",
                 "Numeric date from SDTM.DS.DSSTDTC where DSDECOD='RANDOMIZED'."),
    "TRTSDT":   ("Date of First Exposure to Treatment", "date", "Derived",
                 "Numeric date from SDTM.DM.RFXSTDTC. Defines analysis Day 1."),
    "TRTEDT":   ("Date of Last Exposure to Treatment", "date", "Derived",
                 "Numeric date from SDTM.DM.RFXENDTC."),
    "TRTDURD":  ("Total Treatment Duration (Days)", "integer", "Derived",
                 "TRTEDT - TRTSDT + 1 (both endpoints inclusive)."),
    "EOSSTT":   ("End of Study Status", "text", "Derived",
                 "'COMPLETED' when the DS disposition event is COMPLETED, else 'DISCONTINUED'."),
    "DCSREAS":  ("Reason for Discontinuation from Study", "text", "Derived",
                 "Null when completed; otherwise the disposition event's DSDECOD."),
    "DCSREASP": ("Reason Spec for Discont from Study", "text", "Derived",
                 "Null when completed; otherwise the verbatim disposition DSTERM."),
    "SAFFL":    ("Safety Population Flag", "text", "Derived",
                 "'Y' when the subject took at least one dose (TRTSDT non-missing)."),
    "ITTFL":    ("Intent-To-Treat Population Flag", "text", "Derived",
                 "'Y' when the subject was randomised (RANDDT non-missing)."),
    "COMPLFL":  ("Completers Population Flag", "text", "Derived",
                 "'Y' when EOSSTT='COMPLETED'."),
    "HEIGHTBL": ("Baseline Height (cm)", "float", "Derived",
                 "Last non-missing VS height on or before TRTSDT."),
    "WEIGHTBL": ("Baseline Weight (kg)", "float", "Derived",
                 "Last non-missing VS weight on or before TRTSDT."),
    "BMIBL":    ("Baseline Body Mass Index (kg/m2)", "float", "Derived",
                 "WEIGHTBL / (HEIGHTBL/100)**2, rounded to 0.01."),

    # OCCDS / event predecessors
    "AESEQ":    ("Sequence Number", "integer", "Predecessor", "SDTM.AE.AESEQ."),
    "AETERM":   ("Reported Term for the Adverse Event", "text", "Predecessor", "SDTM.AE.AETERM."),
    "AEDECOD":  ("Dictionary-Derived Term", "text", "Predecessor",
                 "SDTM.AE.AEDECOD (MedDRA-coded; illustrative only)."),
    "AESEV":    ("Severity/Intensity", "text", "Predecessor", "SDTM.AE.AESEV."),
    "AESER":    ("Serious Event", "text", "Predecessor", "SDTM.AE.AESER."),
    "AEREL":    ("Causality", "text", "Predecessor", "SDTM.AE.AEREL."),
    "AEOUT":    ("Outcome of Adverse Event", "text", "Predecessor", "SDTM.AE.AEOUT."),
    "ASEVN":    ("Analysis Severity (N)", "integer", "Derived",
                 "Numeric severity: MILD=1, MODERATE=2, SEVERE=3."),
    "AREL":     ("Analysis Causality", "text", "Derived",
                 "'Y' when AEREL in (RELATED, POSSIBLY RELATED), else 'N'. Collapse per the SAP."),
    "ASTDT":    ("Analysis Start Date", "date", "Derived", "Numeric date from SDTM.AE.AESTDTC."),
    "AENDT":    ("Analysis End Date", "date", "Derived", "Numeric date from SDTM.AE.AEENDTC."),
    "ASTDY":    ("Analysis Start Relative Day", "integer", "Derived",
                 "Study day of ASTDT relative to TRTSDT. No Day 0."),
    "AENDY":    ("Analysis End Relative Day", "integer", "Derived",
                 "Study day of AENDT relative to TRTSDT. No Day 0."),
    "ADURN":    ("Analysis Duration (Days)", "integer", "Derived",
                 "AENDT - ASTDT + 1. Missing (not 0) for an ongoing event."),
    "TRTEMFL":  ("Treatment Emergent Analysis Flag", "text", "Predecessor",
                 "SDTM.SUPPAE.QVAL where QNAM='AETRTEM' — read, not re-derived."),
    "AOCCFL":   ("1st Occurrence of Any AE Flag", "text", "Derived",
                 "'Y' on the first treatment-emergent event per subject (order ASTDT, AESEQ)."),
    "AOCCPFL":  ("1st Occurrence of Preferred Term Flag", "text", "Derived",
                 "'Y' on the first treatment-emergent event per subject per AEDECOD."),

    # BDS common
    "PARAMCD":  ("Parameter Code", "text", "Assigned", "See value-level metadata."),
    "PARAM":    ("Parameter", "text", "Assigned", "See value-level metadata."),
    "PARAMN":   ("Parameter (N)", "integer", "Assigned", "Numeric code for PARAM; see value-level metadata."),
    "PARCAT1":  ("Parameter Category 1", "text", "Predecessor", "SDTM.LB.LBCAT."),
    "AVAL":     ("Analysis Value", "float", "Derived", "See value-level metadata."),
    "AVALU":    ("Analysis Value Unit", "text", "Predecessor", "Standardised unit; see value-level metadata."),
    "AVISIT":   ("Analysis Visit", "text", "Derived",
                 "Analysis visit label mapped from VISIT (Screening / Baseline / Week 4)."),
    "AVISITN":  ("Analysis Visit (N)", "integer", "Derived",
                 "Numeric analysis visit: Screening=0, Baseline=1, Week 4=4."),
    "VISIT":    ("Visit Name", "text", "Predecessor", "SDTM findings VISIT."),
    "VISITNUM": ("Visit Number", "integer", "Predecessor", "SDTM findings VISITNUM."),
    "ADT":      ("Analysis Date", "date", "Derived", "Numeric date from the collected --DTC."),
    "ADY":      ("Analysis Relative Day", "integer", "Derived",
                 "Study day of ADT relative to TRTSDT. No Day 0."),
    "ABLFL":    ("Baseline Record Flag", "text", "Derived",
                 "'Y' on the last non-missing record on or before TRTSDT, per subject per parameter."),
    "BASE":     ("Baseline Value", "float", "Derived",
                 "AVAL of the subject+parameter's ABLFL='Y' record."),
    "CHG":      ("Change from Baseline", "float", "Derived",
                 "AVAL - BASE on post-baseline records (rounded to 0.0001). Missing at baseline."),
    "PCHG":     ("Percent Change from Baseline", "float", "Derived",
                 "100 * CHG / BASE on post-baseline records (rounded to 0.0001). Missing when BASE=0."),
    "DTYPE":    ("Derivation Type", "text", "Derived",
                 "Null throughout — every record is observed; no imputed/derived records."),
    "ANL01FL":  ("Analysis Flag 01", "text", "Derived",
                 "'Y' on the baseline record and on-treatment records."),
    "SRCDOM":   ("Source Domain", "text", "Derived", "SDTM domain the record traces to; null for derived parameters."),
    "SRCVAR":   ("Source Variable", "text", "Derived", "SDTM variable the value traces to; null for derived parameters."),
    "SRCSEQ":   ("Source Sequence Number", "integer", "Derived", "SDTM --SEQ the record traces to; null for derived parameters."),

    # ADLB extras
    "ANRLO":    ("Analysis Normal Range Lower Limit", "float", "Predecessor", "SDTM.LB.LBSTNRLO."),
    "ANRHI":    ("Analysis Normal Range Upper Limit", "float", "Predecessor", "SDTM.LB.LBSTNRHI."),
    "ANRIND":   ("Analysis Reference Range Indicator", "text", "Derived",
                 "LOW / NORMAL / HIGH from AVAL against ANRLO–ANRHI."),
    "BNRIND":   ("Baseline Reference Range Indicator", "text", "Derived",
                 "ANRIND of the subject+parameter's baseline record."),
    "SHIFT1":   ("Shift 1", "text", "Derived",
                 "'BNRIND to ANRIND' on post-baseline records; null at baseline."),
    "CRIT1":    ("Analysis Criterion 1", "text", "Derived",
                 "Text of the prespecified criterion; populated only where it applies (ALT)."),
    "CRIT1FL":  ("Criterion 1 Evaluation Result Flag", "text", "Derived",
                 "'Y'/'N' on ALT records (AVAL > ANRHI); null on parameters the criterion does not apply to."),

    # ADTTE extras
    "STARTDT":  ("Time to Event Origin Date", "date", "Derived", "Time origin, set to ADSL.TRTSDT."),
    "AVAL_TTE": ("Analysis Value (Days)", "integer", "Derived", "ADT - STARTDT + 1 (days)."),
    "CNSR":     ("Censor", "integer", "Derived",
                 "0 = event occurred, 1 = censored. Note: reversed from other ADaM flags."),
    "EVNTDESC": ("Event or Censoring Description", "text", "Derived",
                 "What ADT represents: the event, or the reason for censoring."),
    "CNSDTDSC": ("Censoring Description", "text", "Derived",
                 "Why censored; populated on censored records only."),
}

# Per-dataset variable order. Names resolve against LIB.
ORDER = {
    "ADSL": ["STUDYID", "USUBJID", "SUBJID", "SITEID", "COUNTRY", "ARM", "ACTARM",
             "TRT01P", "TRT01PN", "TRT01A", "TRT01AN", "AGE", "AGEU", "AGEGR1", "AGEGR1N",
             "SEX", "RACE", "ETHNIC", "RFICDT", "RANDDT", "TRTSDT", "TRTEDT", "TRTDURD",
             "EOSSTT", "DCSREAS", "DCSREASP", "SAFFL", "ITTFL", "COMPLFL",
             "HEIGHTBL", "WEIGHTBL", "BMIBL"],
    "ADAE": ["STUDYID", "USUBJID", "SUBJID", "SITEID", "TRTA", "TRTAN", "AGE", "AGEGR1",
             "AGEGR1N", "SEX", "RACE", "SAFFL", "TRTSDT", "TRTEDT", "AESEQ", "AETERM",
             "AEDECOD", "AESEV", "ASEVN", "AESER", "AEREL", "AREL", "AEOUT", "ASTDT",
             "AENDT", "ASTDY", "AENDY", "ADURN", "TRTEMFL", "AOCCFL", "AOCCPFL"],
    "ADVS": ["STUDYID", "USUBJID", "SUBJID", "SITEID", "TRTP", "TRTPN", "TRTA", "TRTAN",
             "AGE", "AGEGR1", "AGEGR1N", "SEX", "RACE", "SAFFL", "ITTFL", "TRTSDT", "TRTEDT",
             "PARAMCD", "PARAM", "PARAMN", "AVAL", "AVALU", "AVISIT", "AVISITN", "VISIT",
             "VISITNUM", "ADT", "ADY", "ABLFL", "BASE", "CHG", "PCHG", "DTYPE", "ANL01FL",
             "SRCDOM", "SRCVAR", "SRCSEQ"],
    "ADLB": ["STUDYID", "USUBJID", "SUBJID", "SITEID", "TRTP", "TRTPN", "TRTA", "TRTAN",
             "AGE", "AGEGR1", "AGEGR1N", "SEX", "RACE", "SAFFL", "ITTFL", "TRTSDT", "TRTEDT",
             "PARAMCD", "PARAM", "PARAMN", "PARCAT1", "AVAL", "AVALU", "AVISIT", "AVISITN",
             "VISIT", "VISITNUM", "ADT", "ADY", "ABLFL", "BASE", "CHG", "PCHG", "DTYPE",
             "ANL01FL", "ANRLO", "ANRHI", "ANRIND", "BNRIND", "SHIFT1", "CRIT1", "CRIT1FL",
             "SRCDOM", "SRCVAR", "SRCSEQ"],
    "ADTTE": ["STUDYID", "USUBJID", "SUBJID", "SITEID", "TRTP", "TRTPN", "TRTA", "TRTAN",
              "AGE", "AGEGR1", "SEX", "RACE", "SAFFL", "ITTFL", "PARAMCD", "PARAM", "PARAMN",
              "STARTDT", "ADT", "AVAL", "CNSR", "EVNTDESC", "CNSDTDSC", "SRCDOM", "SRCVAR"],
}

# ADTTE AVAL uses the TTE derivation, not "see value-level".
LIB_OVERRIDE = {("ADTTE", "AVAL"): LIB["AVAL_TTE"]}

# variable -> codelist OID (applied in define.xml and shown in the spec)
VAR_CODELIST = {
    "SEX": "SEX", "AGEGR1": "AGEGR1", "AGEGR1N": "AGEGR1N",
    "TRT01PN": "TRTN", "TRT01AN": "TRTN", "TRTPN": "TRTN", "TRTAN": "TRTN",
    "SAFFL": "NY", "ITTFL": "NY", "COMPLFL": "NY", "TRTEMFL": "NY",
    "AOCCFL": "NY", "AOCCPFL": "NY", "AREL": "NY", "AESER": "NY", "ANL01FL": "NY",
    "ABLFL": "NY", "CRIT1FL": "NY",
    "AESEV": "AESEV", "ASEVN": "ASEVN", "AEOUT": "AEOUT",
    "ANRIND": "NRIND", "BNRIND": "NRIND", "CNSR": "CNSR",
}

# Value-level metadata: dataset -> list of (PARAMCD, PARAM, PARAMN, unit, AVAL derivation)
VALUELEVEL = {
    "ADVS": [
        ("SYSBP", "Systolic Blood Pressure (mmHg)", 1, "mmHg", "VS.VSSTRESN where VSTESTCD='SYSBP'."),
        ("DIABP", "Diastolic Blood Pressure (mmHg)", 2, "mmHg", "VS.VSSTRESN where VSTESTCD='DIABP'."),
        ("PULSE", "Pulse Rate (beats/min)", 3, "beats/min", "VS.VSSTRESN where VSTESTCD='PULSE'."),
        ("TEMP", "Temperature (C)", 4, "C", "VS.VSSTRESN where VSTESTCD='TEMP'."),
        ("WEIGHT", "Weight (kg)", 5, "kg", "VS.VSSTRESN where VSTESTCD='WEIGHT'."),
        ("HEIGHT", "Height (cm)", 6, "cm", "VS.VSSTRESN where VSTESTCD='HEIGHT' (Screening only)."),
        ("BMI", "Body Mass Index (kg/m2)", 7, "kg/m2",
         "Derived parameter: visit WEIGHT / (ADSL.HEIGHTBL/100)**2, rounded to 0.01. "
         "No SDTM source record (SRCDOM/SRCVAR/SRCSEQ null)."),
    ],
    "ADLB": [
        ("HGB", "Hemoglobin (g/dL)", 1, "g/dL", "LB.LBSTRESN where LBTESTCD='HGB'."),
        ("HCT", "Hematocrit (%)", 2, "%", "LB.LBSTRESN where LBTESTCD='HCT'."),
        ("WBC", "Leukocytes (10^9/L)", 3, "10^9/L", "LB.LBSTRESN where LBTESTCD='WBC'."),
        ("PLAT", "Platelets (10^9/L)", 4, "10^9/L", "LB.LBSTRESN where LBTESTCD='PLAT'."),
        ("ALT", "Alanine Aminotransferase (U/L)", 5, "U/L", "LB.LBSTRESN where LBTESTCD='ALT'."),
        ("CREAT", "Creatinine (mg/dL)", 6, "mg/dL", "LB.LBSTRESN where LBTESTCD='CREAT'."),
    ],
    "ADTTE": [
        ("TTFAE", "Time to First Treatment-Emergent Adverse Event (days)", 1, "days",
         "Days from TRTSDT to the first treatment-emergent AE (ADAE.ASTDT where AOCCFL='Y'); "
         "subjects with no event are censored at TRTEDT."),
    ],
}

# Codelists the datasets apply.
CODELISTS = {
    "NY":     ("No Yes Response", [("N", "No"), ("Y", "Yes")]),
    "SEX":    ("Sex", [("M", "Male"), ("F", "Female"), ("U", "Unknown")]),
    "AGEGR1": ("Pooled Age Group 1", [("<65", "Less than 65 years"), (">=65", "65 years or older")]),
    "AGEGR1N": ("Pooled Age Group 1 (N)", [("1", "<65"), ("2", ">=65")]),
    "TRTN":   ("Treatment (N)", [("1", "Placebo"), ("2", "Drug A")]),
    "AESEV":  ("Severity/Intensity Scale", [("MILD", "Mild"), ("MODERATE", "Moderate"),
                                            ("SEVERE", "Severe")]),
    "ASEVN":  ("Analysis Severity (N)", [("1", "Mild"), ("2", "Moderate"), ("3", "Severe")]),
    "AEOUT":  ("Outcome of Event", [("RECOVERED/RESOLVED", "Recovered/Resolved"),
                                    ("RECOVERING/RESOLVING", "Recovering/Resolving"),
                                    ("NOT RECOVERED/NOT RESOLVED", "Not Recovered/Not Resolved")]),
    "NRIND":  ("Reference Range Indicator", [("LOW", "Low"), ("NORMAL", "Normal"),
                                             ("HIGH", "High")]),
    "CNSR":   ("Censor", [("0", "Event"), ("1", "Censored")]),
}


def var_meta(ds, name):
    return LIB_OVERRIDE.get((ds, name)) or LIB[name]


# ---------------------------------------------------------------- workbook

INK = "0F2E3D"
HEADFILL = PatternFill("solid", fgColor=INK)
HEADFONT = Font(color="FFFFFF", bold=True, size=11)
TITLEFONT = Font(color=INK, bold=True, size=14)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="CFDEE1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADFILL
        cell.font = HEADFONT
        cell.alignment = WRAP
        cell.border = BORDER


def build():
    wb = openpyxl.Workbook()

    # --- Index sheet ---
    ws = wb.active
    ws.title = "Index"
    ws["A1"] = "ADaM Specification — Study ABC-01 (synthetic)"
    ws["A1"].font = TITLEFONT
    ws["A2"] = "Standard: ADaM-IG v1.2 (ADaM v2.1).  All data synthetic; coded terms illustrative."
    ws["A2"].font = Font(italic=True, color="5A7682")
    hdr = ["Dataset", "Label", "Class", "Structure", "Key Variables"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    for ds, (label, cls, structure, key) in DATASETS.items():
        ws.append([ds, label, cls, structure, key])
    style_header(ws, hrow, len(hdr))
    for r in range(hrow + 1, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
    for col, w in zip("ABCDE", (9, 34, 26, 46, 34)):
        ws.column_dimensions[col].width = w

    # --- one sheet per dataset ---
    for ds in DATASETS:
        ws = wb.create_sheet(ds)
        ws["A1"] = f"{ds} — {DATASETS[ds][0]}"
        ws["A1"].font = TITLEFONT
        ws["A2"] = (f"Class: {DATASETS[ds][1]}   |   Structure: {DATASETS[ds][2]}   |   "
                    f"Key: {DATASETS[ds][3]}")
        ws["A2"].font = Font(italic=True, color="5A7682")
        hdr = ["#", "Variable", "Label", "Type", "Origin", "Codelist", "Derivation / Source"]
        ws.append([])
        ws.append(hdr)
        hrow = ws.max_row
        for i, name in enumerate(ORDER[ds], start=1):
            label, dtype, origin, deriv = var_meta(ds, name)
            cl = VAR_CODELIST.get(name, "")
            ws.append([i, name, label, dtype, origin, cl, deriv])
        style_header(ws, hrow, len(hdr))
        for r in range(hrow + 1, ws.max_row + 1):
            for c in range(1, len(hdr) + 1):
                ws.cell(row=r, column=c).alignment = WRAP
                ws.cell(row=r, column=c).border = BORDER
        for col, w in zip("ABCDEFG", (4, 12, 34, 9, 13, 10, 60)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    # --- Value-level metadata ---
    ws = wb.create_sheet("ValueLevel")
    ws["A1"] = "Value-Level Metadata — parameter-specific AVAL derivations"
    ws["A1"].font = TITLEFONT
    ws["A2"] = ("The BDS/TTE payload. Each PARAMCD names its own AVAL source, so a reviewer can "
                "trace one parameter without reading the others.")
    ws["A2"].font = Font(italic=True, color="5A7682")
    hdr = ["Dataset", "PARAMCD", "PARAM", "PARAMN", "Unit (AVALU)", "AVAL Derivation"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    for ds, items in VALUELEVEL.items():
        for pc, pm, pn, unit, deriv in items:
            ws.append([ds, pc, pm, pn, unit, deriv])
    style_header(ws, hrow, len(hdr))
    for r in range(hrow + 1, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = WRAP
            ws.cell(row=r, column=c).border = BORDER
    for col, w in zip("ABCDEF", (9, 10, 42, 9, 12, 60)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    # --- Codelists ---
    ws = wb.create_sheet("Codelists")
    ws["A1"] = "Controlled Terminology applied"
    ws["A1"].font = TITLEFONT
    hdr = ["Codelist", "Name", "Coded Value", "Decode"]
    ws.append([])
    ws.append(hdr)
    hrow = ws.max_row
    for oid, (name, items) in CODELISTS.items():
        for j, (code, decode) in enumerate(items):
            ws.append([oid if j == 0 else "", name if j == 0 else "", code, decode])
    style_header(ws, hrow, len(hdr))
    for r in range(hrow + 1, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).alignment = TOP
            ws.cell(row=r, column=c).border = BORDER
    for col, w in zip("ABCD", (12, 34, 28, 30)):
        ws.column_dimensions[col].width = w

    return wb


def main():
    wb = build()
    wb.save(OUT)
    nvars = sum(len(v) for v in ORDER.values())
    nvlm = sum(len(v) for v in VALUELEVEL.values())
    print(f"wrote {os.path.basename(OUT)}")
    print(f"  datasets:            {len(DATASETS)}")
    print(f"  variables:           {nvars}")
    print(f"  value-level rows:    {nvlm}")
    print(f"  codelists:           {len(CODELISTS)}")

    # self-check: every ORDER name resolves, every ADaM CSV column is specified.
    import csv
    for ds in DATASETS:
        for name in ORDER[ds]:
            assert var_meta(ds, name), f"{ds}.{name} has no metadata"
        path = os.path.join(HERE, "adam", f"{ds.lower()}.csv")
        if os.path.exists(path):
            cols = list(next(csv.reader(open(path))))
            spec_cols = set(ORDER[ds])
            missing = [c for c in cols if c not in spec_cols]
            extra = [c for c in ORDER[ds] if c not in cols]
            assert not missing, f"{ds}: columns in data but not in spec: {missing}"
            assert not extra, f"{ds}: variables in spec but not in data: {extra}"
    print("  self-check: spec covers every column of every ADaM dataset exactly")


if __name__ == "__main__":
    main()
